"""
Centralized UI Utilities for IBI Reporting Application.
Provides consistent styling, download with MongoDB logging, and date-stamped filenames.
Professional-grade utilities for enterprise deployment.
"""

import streamlit as st
from datetime import datetime
from typing import Optional, Dict, Any, List, Union
import pandas as pd
from io import BytesIO
import os
import sys
import logging

# Configure logging
logger = logging.getLogger(__name__)

# Import MongoDB utilities from common.mongo
try:
    import common.mongo as mongo
    
    # We use a property-like check to ensure we get the CURRENT connection status
    def is_mongo_available():
        return getattr(mongo, "MONGO_CONNECTED", False)
        
    log_report_download = getattr(mongo, "log_report_download", None)
    log_multi_report_download = getattr(mongo, "log_multi_report_download", None)
    save_module_report = getattr(mongo, "save_module_report", None)
    save_and_track_report = getattr(mongo, "save_and_track_report", None)
    save_reconciliation_report = getattr(mongo, "save_reconciliation_report", None)
    save_report_with_tracking = getattr(mongo, "save_report_with_tracking", None)
    register_report_info = getattr(mongo, "register_report_info", None)
    
except ImportError:
    def is_mongo_available():
        return False
    log_report_download = None
    log_multi_report_download = None
    save_module_report = None
    save_and_track_report = None
    save_reconciliation_report = None
    save_report_with_tracking = None
    register_report_info = None


def apply_professional_style():
    """Finalized ultra-aggressive UI styling to force professional LIGHT mode across all tools."""
    st.markdown("""
        <style>
        /* ======== FORCE LIGHT MODE (ULTRA) ======== */
        :root {
            color-scheme: light !important;
            --primary-color: #1e40af !important;
            --background-color: #ffffff !important;
            --secondary-background-color: #f8fafc !important;
            --text-color: #1f2937 !important;
        }
        
        /* Force app background and all text headings/paragraphs */
        .stApp, [data-testid="stAppViewContainer"], .main {
            background-color: #ffffff !important;
        }
        
        .stApp p, .stApp span, .stApp label, .stApp div, .stApp li {
            color: #1f2937 !important;
        }
        
        h1, h2, h3, h4, h5, h6 {
            color: #1e3a8a !important;
            font-family: 'Inter', 'Segoe UI', sans-serif !important;
        }

        /* ======== SIDEBAR STYLING ======== */
        section[data-testid="stSidebar"] {
            background-color: #f8fafc !important;
            border-right: 1px solid #e2e8f0 !important;
            min-width: 330px !important;
        }
        
        [data-testid="stSidebar"] .st-emotion-cache-1kyx73u, 
        [data-testid="stSidebar"] [data-testid="stVerticalBlock"],
        [data-testid="stSidebarNav"] {
            background-color: #f8fafc !important;
        }

        /* Hide Streamlit elements (BUT keep collapsedControl for sidebar toggle) */
        /* Hide standard Streamlit header elements to keep it clean */
        [data-testid="stHeaderSettingsButton"], .stDeployButton, [data-testid="stAppViewMenu"], footer {
            display: none !important;
        }
        
        [data-testid="stHeader"] {
            background: transparent !important;
            border: none !important;
            z-index: 99 !important;
        }

        /* FORCE Sidebar Visibility and make it static */
        section[data-testid="stSidebar"] {
            display: flex !important;
            visibility: visible !important;
            z-index: 100 !important;
        }

        /* Pinned Toggle Button - ensuring it's always accessible at top-left */
        [data-testid="collapsedControl"], 
        button[data-testid="stSidebarCollapseButton"] {
            position: fixed !important;
            top: 5px !important;
            left: 5px !important;
            z-index: 1000000 !important;
            display: flex !important;
            visibility: visible !important;
            background: white !important;
            border: 2px solid #1e40af !important;
            border-radius: 8px !important;
            box-shadow: 0 4px 10px rgba(0,0,0,0.2) !important;
            padding: 4px !important;
            opacity: 1 !important;
        }
        
        [data-testid="collapsedControl"] svg,
        button[data-testid="stSidebarCollapseButton"] svg {
            fill: #1e40af !important;
        }

        /* Ensure sidebar content doesn't get cut off */
        [data-testid="stSidebar"] [data-testid="stSidebarHeader"] {
            padding-top: 50px !important;
        }

        /* Adjust sidebar width to be more container-friendly */
        section[data-testid="stSidebar"] {
            min-width: 300px !important;
            max-width: 350px !important;
        }

        /* ======== FORM ELEMENTS ======== */
        /* Universal background reset for all inputs to white */
        div[data-baseweb="select"] > div,
        div[data-baseweb="input"] > div,
        div[data-baseweb="textarea"] > div,
        div[data-baseweb="base-input"] > div,
        .stTextInput input, .stNumberInput input, .stTextArea textarea {
            background-color: #ffffff !important;
            color: #1f2937 !important;
            border-color: #cbd5e1 !important;
            border-radius: 8px !important;
        }

        /* Dropdowns, Popovers, Menus */
        div[data-baseweb="popover"], div[data-baseweb="menu"], ul[role="listbox"], [data-testid="stTooltipHoverTarget"] {
            background-color: #ffffff !important;
            color: #1f2937 !important;
            box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.1) !important;
        }
        
        li[role="option"], [data-baseweb="select"] li {
            background-color: #ffffff !important;
            color: #1f2937 !important;
        }
        
        li[role="option"]:hover, [data-baseweb="select"] li:hover {
            background-color: #eff6ff !important;
            color: #1e40af !important;
        }

        /* Labels and Descriptions */
        label, .st-at, .st-ae, .st-af, [data-testid="stMarkdownContainer"] p {
            color: #334155 !important;
        }

        /* File Uploader styling */
        [data-testid="stFileUploader"] {
            border: 2px dashed #3b82f6 !important;
            background-color: #f8fafc !important;
            padding: 2rem !important;
            border-radius: 12px !important;
        }
        
        [data-testid="stFileUploader"] section {
            background-color: transparent !important;
        }

        [data-testid="stFileUploader"] button {
            background-color: #1e40af !important;
            color: white !important;
            border-radius: 8px !important;
        }

        /* ======== PREMIUM DASHBOARD ELEMENTS ======== */
        .stButton button {
            border-radius: 8px !important;
            background: #1e40af !important;
            color: white !important;
            font-weight: 600 !important;
            border: none !important;
            padding: 0.5rem 2rem !important;
            transition: all 0.2s ease !important;
        }

        .stButton button:hover {
            background: #1d4ed8 !important;
            transform: translateY(-1px);
            box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1) !important;
        }

        /* Tabs */
        .stTabs [data-baseweb="tab-list"] {
            background-color: #f1f5f9 !important;
            border-radius: 10px;
            padding: 5px;
        }
        
        .stTabs [data-baseweb="tab"] {
            color: #64748b !important;
        }
        
        .stTabs [aria-selected="true"] {
            background-color: #ffffff !important;
            color: #1e40af !important;
            font-weight: 700 !important;
            border-radius: 8px !important;
        }

        /* Metrics - Keep the delta colors but ensure labels are dark */
        [data-testid="stMetricValue"] {
            color: #1e40af !important;
        }
        
        [data-testid="stMetricLabel"] p {
            color: #475569 !important;
        }

        /* Dataframes */
        [data-testid="stDataFrame"], [data-testid="stTable"] {
            background-color: #ffffff !important;
            border-radius: 10px !important;
            border: 1px solid #e2e8f0 !important;
        }

        /* Fix for scrollbars in dark mode browsers */
        ::-webkit-scrollbar {
            width: 10px;
            height: 10px;
        }
        ::-webkit-scrollbar-track {
            background: #f1f5f9;
        }
        ::-webkit-scrollbar-thumb {
            background: #cbd5e1;
            border-radius: 5px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background: #94a3b8;
        }
        </style>
    """, unsafe_allow_html=True)


def get_download_filename(base_name: str, extension: str = "xlsx") -> str:
    """
    Generates a filename with exact current date and time.
    Format: base_name_2026-01-11_15-30-45.xlsx
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base_name = base_name.replace(f".{extension}", "").replace(" ", "_")
    # Clean up any double underscores
    base_name = "_".join(filter(None, base_name.split("_")))
    return f"{base_name}_{timestamp}.{extension}"


def render_header(title: str, subtitle: Optional[str] = None):
    """Renders a styled header section."""
    st.markdown(f"""
        <div class="report-header">
            <h2 class="report-title">{title}</h2>
            {f'<p style="color: #94a3b8;">{subtitle}</p>' if subtitle else ''}
        </div>
    """, unsafe_allow_html=True)


def to_excel(df: pd.DataFrame, apply_doc_formatting: bool = False, sheet_name: str = 'Report') -> bytes:
    """Convert DataFrame to Excel bytes with optional formatting."""
    output = BytesIO()
    # Handle MultiIndex or standard Index
    index_needed = isinstance(df.index, pd.MultiIndex) or (df.index.name is not None)
    
    # Sanitize sheet name - remove invalid Excel characters
    import re
    sheet_name = re.sub(r'[\\/\*\?\:\[\]]', '_', str(sheet_name))[:31]
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=index_needed, sheet_name=sheet_name)
        
        if apply_doc_formatting and 'DOC' in df.columns:
            try:
                from openpyxl.styles import PatternFill, Font
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                doc_col_idx = df.columns.get_loc('DOC') + (2 if index_needed else 1)
                
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=doc_col_idx)
                    doc_value = df.iloc[row_idx - 2]['DOC']
                    
                    if pd.notna(doc_value) and doc_value != '':
                        try:
                            doc_val = float(doc_value)
                            if doc_val < 7:
                                color = 'FF4444'
                            elif doc_val < 15:
                                color = 'FF8800'
                            elif doc_val < 30:
                                color = '44FF44'
                            elif doc_val < 45:
                                color = 'FFFF44'
                            elif doc_val < 60:
                                color = '44DDFF'
                            elif doc_val < 90:
                                color = '8B4513'
                            else:
                                color = '000000'
                            
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            cell.font = Font(color='FFFFFF' if doc_val < 15 or doc_val >= 60 else '000000', bold=True)
                        except (ValueError, TypeError):
                            pass
            except ImportError:
                pass
    
    return output.getvalue()


def to_multi_sheet_excel(reports: Dict[str, pd.DataFrame]) -> bytes:
    """
    Convert multiple DataFrames to a multi-sheet Excel file.
    
    Args:
        reports: Dict of {sheet_name: DataFrame}
    
    Returns:
        Excel file as bytes
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in reports.items():
            # Clean sheet name (Excel has 31 char limit and special char restrictions)
            clean_name = str(sheet_name)[:31].replace('/', '-').replace('\\', '-')
            clean_name = clean_name.replace('[', '').replace(']', '').replace('*', '')
            clean_name = clean_name.replace('?', '').replace(':', '-')
            
            # Handle MultiIndex
            index_needed = isinstance(df.index, pd.MultiIndex) or (df.index.name is not None)
            df.to_excel(writer, index=index_needed, sheet_name=clean_name)
    
    return output.getvalue()


def _process_pending_mongo_logs():
    """Process any pending MongoDB logs from previous download clicks."""
    if "pending_mongo_logs" not in st.session_state:
        st.session_state.pending_mongo_logs = []
        return
    
    pending = st.session_state.pending_mongo_logs
    if not pending:
        return
    
    # Process all pending logs
    for log_entry in pending:
        try:
            if log_report_download:
                success = log_report_download(
                    user_email=log_entry.get("user", "anonymous"),
                    module=log_entry.get("module", "unknown"),
                    report_name=log_entry.get("report_name", "Report"),
                    filename=log_entry.get("filename", "report.xlsx"),
                    df_data=log_entry.get("df_data"),
                    row_count=log_entry.get("row_count", 0),
                    col_count=log_entry.get("col_count", 0),
                    file_size=log_entry.get("file_size", 0),
                    sheet_name=log_entry.get("sheet_name")
                )
                if success:
                    logger.info(f"✅ Logged to MongoDB: {log_entry.get('report_name')}")
        except Exception as e:
            logger.warning(f"MongoDB log error: {e}")
    
    # Clear processed logs
    st.session_state.pending_mongo_logs = []


def _queue_mongo_log(df: pd.DataFrame, filename: str, module_name: str, 
                     report_name: str, file_size: int, sheet_name: str = None):
    """Queue a MongoDB log entry to be processed on next rerun."""
    if "pending_mongo_logs" not in st.session_state:
        st.session_state.pending_mongo_logs = []
    
    user = st.session_state.get("user", "anonymous")
    
    log_entry = {
        "user": user,
        "module": module_name,
        "report_name": report_name,
        "filename": filename,
        "df_data": df,
        "row_count": len(df) if hasattr(df, '__len__') else 0,
        "col_count": len(df.columns) if hasattr(df, 'columns') else 0,
        "file_size": file_size,
        "sheet_name": sheet_name
    }
    
    st.session_state.pending_mongo_logs.append(log_entry)
    logger.info(f"📋 Queued for MongoDB: {report_name}")


# =============================================================================
# AUTO-SAVE GENERATED REPORTS (NEW - Immediate save on generation)
# =============================================================================

def _get_saved_reports_key(module_name: str) -> str:
    """Get session state key for tracking saved reports."""
    return f"_saved_reports_{module_name}"


def auto_save_generated_reports(reports: Dict[str, pd.DataFrame], module_name: str,
                                 show_toast: bool = True, tool_name: str = None,
                                 metadata: Dict = None) -> int:
    """
    AUTO-SAVE all generated reports to MongoDB immediately when called.
    
    Args:
        reports: Dict of {report_name: DataFrame}
        module_name: Module name (e.g., 'amazon')
        show_toast: Whether to show success toast
        tool_name: Tool name (auto-detected if not provided)
        metadata: Optional additional metadata to include
    """
    if not is_mongo_available() or not save_report_with_tracking:
        logger.warning("MongoDB not available for auto-save")
        return 0
    
    # Auto-detect tool name from calling file if not provided
    if tool_name is None:
        try:
            import inspect
            import os
            frame = inspect.currentframe()
            caller_frame = frame.f_back
            if caller_frame:
                caller_file = caller_frame.f_code.co_filename
                tool_name = os.path.basename(caller_file).replace('.py', '').replace('_', ' ').title()
        except:
            tool_name = "Unknown Tool"
            
    # Track which reports have been saved this session to avoid duplicates
    saved_key = _get_saved_reports_key(module_name)
    if saved_key not in st.session_state:
        st.session_state[saved_key] = set()
    
    user = st.session_state.get("user", "anonymous")
    success_count = 0
    new_saves = []
    
    # Prepare base metadata
    base_metadata = {
        "auto_saved": True,
        "generated_at": datetime.now().isoformat()
    }
    if metadata:
        base_metadata.update(metadata)
    
    for report_name, df in reports.items():
        # Handle cases where df might be a list or other object
        row_count = len(df) if hasattr(df, '__len__') else 0
        
        # Skip if already saved this session
        report_hash = f"{report_name}_{row_count}"
        if report_hash in st.session_state[saved_key]:
            continue
        
        try:
            filename = f"auto_{get_download_filename(report_name.replace(' ', '_'))}"
            
            success = save_report_with_tracking(
                module_name=module_name,
                tool_name=tool_name,
                report_name=report_name,
                df_data=df,
                user_email=user,
                filename=filename,
                metadata=base_metadata
            )
            
            if success:
                success_count += 1
                new_saves.append(report_name)
                st.session_state[saved_key].add(report_hash)
                logger.info(f"✅ Auto-saved: {report_name} from {tool_name} ({module_name})")
                
        except Exception as e:
            logger.warning(f"Auto-save error for {report_name}: {e}")
    
    if new_saves and show_toast:
        st.toast(f"💾 {len(new_saves)} reports saved to database", icon="✅")
    
    return success_count


def _process_pending_download_history():
    """Process any pending download history logs from previous clicks."""
    if "pending_download_history" not in st.session_state:
        st.session_state.pending_download_history = []
        return
    
    pending = st.session_state.pending_download_history
    if not pending:
        return
    
    try:
        from common.mongo import db as mongo_db
        if mongo_db is None:
            return
        
        for record in pending:
            try:
                mongo_db.download_history.insert_one(record)
                logger.info(f"📥 Download history saved: {record.get('report_name')}")
            except Exception as e:
                logger.warning(f"Error saving download history: {e}")
        
        # Clear processed logs
        st.session_state.pending_download_history = []
    except Exception as e:
        logger.warning(f"Process pending download history error: {e}")


def log_download_event(module_name: str, report_name: str, filename: str, tool_name: str = None):
    """
    Queue download event for logging to download_history collection.
    
    Args:
        module_name: Module name
        report_name: Report name
        filename: Downloaded filename
        tool_name: Name of the tool
    """
    if "pending_download_history" not in st.session_state:
        st.session_state.pending_download_history = []
    
    user = st.session_state.get("user", "anonymous")
    
    # Create lightweight download record
    download_record = {
        "user_email": user,
        "report_name": report_name,
        "tool_name": tool_name or "Unknown Tool",
        "module_name": module_name,
        "filename": filename,
        "downloaded_at": datetime.now()
    }
    
    # Queue for processing on next rerun
    st.session_state.pending_download_history.append(download_record)
    logger.info(f"📋 Queued download history: {report_name}")
    
    # Also try immediate save
    if is_mongo_available():
        try:
            from common.mongo import history_col
            if history_col is not None:
                result = history_col.insert_one(download_record.copy())
                if result.inserted_id:
                    st.toast(f"✅ Download logged: {report_name}", icon="📊")
                    logger.info(f"📥 Immediate download history save: {report_name}")
            else:
                # Try getting it again
                from common.mongo import get_download_history_collection
                col = get_download_history_collection()
                if col is not None:
                    col.insert_one(download_record.copy())
                    st.toast(f"✅ Download logged: {report_name}", icon="📊")
        except Exception as e:
            logger.warning(f"Immediate download history error: {e}")
            st.toast(f"⚠️ Could not log download: {str(e)[:50]}", icon="⚠️")



def download_report(df: pd.DataFrame, base_filename: str, button_label: str = "📥 Download Report",
                    module_name: str = "unknown", report_name: str = "Report",
                    apply_doc_formatting: bool = False, key: str = None,
                    sheet_name: str = "Report") -> bool:
    """
    Creates download button and logs full report data to MongoDB when downloaded.
    Uses session state to ensure logging happens even with Streamlit reruns.
    
    Args:
        df: DataFrame to download
        base_filename: Base name for the file (timestamp will be added)
        button_label: Label for the download button
        module_name: Name of the module (for MongoDB logging)
        report_name: Human-readable report name (for MongoDB logging)
        apply_doc_formatting: Whether to apply DOC column formatting
        key: Unique key for the button
        sheet_name: Sheet name for Excel file
    
    Returns:
        True if download was clicked, False otherwise
    """
    # Process any pending logs from previous clicks
    _process_pending_mongo_logs()
    
    # Generate filename with timestamp
    filename = get_download_filename(base_filename)
    
    # Convert to Excel
    excel_data = to_excel(df, apply_doc_formatting, sheet_name)
    
    # Create download button
    btn_key = key or f"download_{base_filename}_{hash(report_name) % 10000}"
    downloaded = st.download_button(
        label=button_label,
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=btn_key
    )
    
    # Queue log for MongoDB when button is clicked (will be processed on rerun)
    if downloaded:
        _queue_mongo_log(df, filename, module_name, report_name, len(excel_data), sheet_name)
        # Also try immediate logging
        _log_download(df, filename, module_name, report_name, len(excel_data), sheet_name)
    
    return downloaded


def download_multi_sheet_excel(reports: Dict[str, pd.DataFrame], base_filename: str,
                                module_name: str, button_label: str = "📥 Download All Reports",
                                key: str = None) -> bool:
    """
    Download multiple reports as a multi-sheet Excel file with MongoDB logging.
    
    Args:
        reports: Dict of {sheet_name: DataFrame}
        base_filename: Base name for the file
        module_name: Name of the module for logging
        button_label: Label for the download button
        key: Unique key for the button
    
    Returns:
        True if download was clicked, False otherwise
    """
    if not reports:
        st.warning("No reports to download")
        return False
    
    # Generate filename with timestamp
    filename = get_download_filename(base_filename)
    
    # Convert to multi-sheet Excel
    excel_data = to_multi_sheet_excel(reports)
    
    # Create download button
    downloaded = st.download_button(
        label=button_label,
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key or f"download_multi_{base_filename}_{datetime.now().timestamp()}"
    )
    
    # Log each report to MongoDB when button is clicked
    if downloaded:
        user = st.session_state.get("user", "anonymous")
        
        if log_multi_report_download:
            try:
                log_multi_report_download(
                    user_email=user,
                    module=module_name,
                    reports=reports,
                    filename=filename,
                    metadata={"multi_sheet": True, "sheet_count": len(reports)}
                )
                st.toast(f"✅ {len(reports)} reports saved to database", icon="📥")
            except Exception as e:
                logger.warning(f"Multi-report log error: {e}")
    
    return downloaded


def _log_download(df: pd.DataFrame, filename: str, module_name: str, 
                  report_name: str, file_size: int, sheet_name: str = None):
    """Internal function to log download to MongoDB."""
    user = st.session_state.get("user", "anonymous")
    
    if log_report_download:
        try:
            success = log_report_download(
                user_email=user,
                module=module_name,
                report_name=report_name,
                filename=filename,
                df_data=df,
                row_count=len(df),
                col_count=len(df.columns),
                file_size=file_size,
                sheet_name=sheet_name
            )
            if success:
                st.toast(f"✅ {report_name} saved to database", icon="📥")
        except Exception as e:
            logger.warning(f"MongoDB log error: {e}")


def auto_log_reports(reports_dict: Dict[str, pd.DataFrame], module_name: str):
    """
    Automatically log multiple reports to MongoDB without waiting for download clicks.
    Useful for tools that generate multiple outputs at once.
    """
    if not is_mongo_available() or not log_report_download:
        return False
    
    user = st.session_state.get("user", "anonymous")
    success_count = 0
    
    for report_name, df in reports_dict.items():
        try:
            filename = f"auto_{get_download_filename(report_name)}"
            success = log_report_download(
                user_email=user,
                module=module_name,
                report_name=report_name,
                filename=filename,
                df_data=df,
                row_count=len(df),
                col_count=len(df.columns)
            )
            if success:
                success_count += 1
        except Exception as e:
            logger.warning(f"Auto-log error for {report_name}: {e}")
            
    if success_count > 0:
        st.toast(f"✅ {success_count} reports saved to MongoDB", icon="💾")
    
    return success_count == len(reports_dict)


def display_dataframe(df: pd.DataFrame, height: int = 400, use_container_width: bool = True):
    """Display DataFrame with proper styling."""
    st.dataframe(df, use_container_width=use_container_width, height=height)


def create_download_section(reports: Dict[str, pd.DataFrame], module_name: str,
                            section_title: str = "📥 Download Reports"):
    """
    Create a standardized download section with individual and combined download options.
    
    Args:
        reports: Dict of {report_name: DataFrame}
        module_name: Module name for logging
        section_title: Title for the download section
    """
    st.markdown(f"### {section_title}")
    
    # Individual downloads
    cols = st.columns(min(len(reports), 3))
    for idx, (report_name, df) in enumerate(reports.items()):
        with cols[idx % 3]:
            download_report(
                df=df,
                base_filename=report_name.replace(" ", "_"),
                button_label=f"📥 {report_name}",
                module_name=module_name,
                report_name=report_name,
                key=f"dl_{module_name}_{report_name}_{idx}"
            )
    
    # Combined download if multiple reports
    if len(reports) > 1:
        st.markdown("---")
        download_multi_sheet_excel(
            reports=reports,
            base_filename=f"{module_name}_all_reports",
            module_name=module_name,
            button_label="📦 Download All Reports (Combined Excel)",
            key=f"dl_combined_{module_name}"
        )


# =============================================================================
# Module-Specific Report Download (NEW STRUCTURE)
# =============================================================================

def download_module_report(df: pd.DataFrame, module_name: str, report_name: str,
                           button_label: str = "📥 Download", key: str = None,
                           apply_doc_formatting: bool = False,
                           tool_name: str = None) -> bool:
    """
    Download button that AUTOMATICALLY saves report to MongoDB when called.
    Also logs download event when user clicks the download button.
    
    Uses session state to avoid duplicate saves on Streamlit reruns.
    
    Args:
        df: DataFrame to download
        module_name: Module/collection name (e.g., 'amazon')
        report_name: Report name (e.g., 'Brand Manager Analysis')
        button_label: Label for download button
        key: Unique key for the button
        apply_doc_formatting: Whether to apply DOC column formatting
        tool_name: Name of the tool that generated this report (auto-detected if not provided)
    
    Returns:
        True if download was clicked
    """
    # Handle Dict input (multi-sheet) - convert to single DataFrame or handle specially
    if isinstance(df, dict):
        # If passed a dict, just use the first one or combine
        if len(df) > 0:
            first_key = list(df.keys())[0]
            df = df[first_key]
        else:
            return False
    
    # Auto-detect tool name from calling file if not provided
    if tool_name is None:
        try:
            import inspect
            import os
            # Get the caller's frame (skip this function)
            frame = inspect.currentframe()
            caller_frame = frame.f_back
            if caller_frame:
                caller_file = caller_frame.f_code.co_filename
                # Extract filename without extension and format nicely
                tool_name = os.path.basename(caller_file).replace('.py', '').replace('_', ' ').title()
        except:
            tool_name = "Unknown Tool"
    
    # Process any pending logs from previous clicks
    _process_pending_mongo_logs()
    _process_pending_download_history()
    
    # ========= AUTO-SAVE: Save to MongoDB when function is called =========
    saved_key = _get_saved_reports_key(module_name)
    if saved_key not in st.session_state:
        st.session_state[saved_key] = set()
    
    # Create unique hash for this report (based on name and row count)
    report_hash = f"{report_name}_{len(df) if hasattr(df, '__len__') else 0}"
    
    # Auto-save if not already saved this session
    if report_hash not in st.session_state[saved_key] and is_mongo_available():
        try:
            user = st.session_state.get("user", "anonymous")
            auto_filename = f"auto_{get_download_filename(report_name.replace(' ', '_'))}"
            
            # Use new save_report_with_tracking for centralized registry tracking
            if save_report_with_tracking:
                report_id = save_report_with_tracking(
                    module_name=module_name,
                    tool_name=tool_name,
                    report_name=report_name,
                    df_data=df,
                    user_email=user,
                    filename=auto_filename,
                    metadata={
                        "auto_saved": True, 
                        "generated_at": datetime.now().isoformat()
                    }
                )
                if report_id:
                    st.session_state[saved_key].add(report_hash)
                    logger.info(f"✅ Auto-saved: {report_name} from {tool_name} ({module_name}) [ID: {report_id}]")
            elif log_report_download:
                # Fallback to old method if new function not available
                success = log_report_download(
                    user_email=user,
                    module=module_name,
                    report_name=report_name,
                    filename=auto_filename,
                    df_data=df,
                    row_count=len(df) if hasattr(df, '__len__') else 0,
                    col_count=len(df.columns) if hasattr(df, 'columns') else 0,
                    metadata={
                        "auto_saved": True, 
                        "generated_at": datetime.now().isoformat(),
                        "tool_name": tool_name
                    }
                )
                if success:
                    st.session_state[saved_key].add(report_hash)
                    logger.info(f"✅ Auto-saved (legacy): {report_name} from {tool_name} ({module_name})")
        except Exception as e:
            logger.warning(f"Auto-save error for {report_name}: {e}")
    # ======================================================================
    
    # Generate filename with timestamp
    base_filename = report_name.replace(" ", "_").lower()
    filename = get_download_filename(base_filename)
    
    # Convert to Excel
    excel_data = to_excel(df, apply_doc_formatting, report_name[:31])
    
    # Create download button with stable key
    btn_key = key or f"dl_{module_name}_{report_name.replace(' ', '_')}_{hash(report_name) % 10000}"
    
    # Initialize download tracking in session state
    if "_download_clicks" not in st.session_state:
        st.session_state._download_clicks = {}
    
    # Check if this button was clicked in a previous run (and not yet logged)
    click_key = f"clicked_{btn_key}"
    if click_key in st.session_state._download_clicks:
        # This was clicked before - log it now (on reload)
        click_info = st.session_state._download_clicks.pop(click_key)
        log_download_event(
            click_info["module"], 
            click_info["report"], 
            click_info["filename"], 
            click_info["tool"]
        )
    
    downloaded = st.download_button(
        label=button_label,
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=btn_key
    )
    
    # When clicked, store info for next reload to log
    if downloaded:
        # Store for next run
        st.session_state._download_clicks[click_key] = {
            "module": module_name,
            "report": report_name,
            "filename": filename,
            "tool": tool_name
        }
        # Also try immediate log
        log_download_event(module_name, report_name, filename, tool_name)
    
    return downloaded


def save_module_reports_on_generate(reports: Dict[str, pd.DataFrame], module_name: str):
    """
    Save all reports to module collection when generated (before download).
    
    Call this after generating reports to pre-save them to the module collection.
    Downloads will then update the download tracking.
    
    Args:
        reports: Dict of {report_name: DataFrame}
        module_name: Module/collection name
    """
    if not is_mongo_available() or not save_module_report:
        return
    
    user = st.session_state.get("user", "anonymous")
    
    for report_name, df in reports.items():
        try:
            save_module_report(
                module_name=module_name,
                report_name=report_name,
                df_data=df,
                user_email=user,
                metadata={
                    "row_count": len(df),
                    "column_count": len(df.columns),
                    "auto_saved": True
                }
            )
        except Exception as e:
            logger.warning(f"Auto-save error for {report_name}: {e}")


def create_module_download_section(reports: Dict[str, pd.DataFrame], module_name: str,
                                    section_title: str = "📥 Download Reports"):
    """
    Create download section with module-specific collection saving.
    
    Args:
        reports: Dict of {report_name: DataFrame}
        module_name: Module/collection name
        section_title: Title for the section
    """
    st.markdown(f"### {section_title}")
    
    cols = st.columns(min(len(reports), 3))
    for idx, (report_name, df) in enumerate(reports.items()):
        with cols[idx % 3]:
            download_module_report(
                df=df,
                module_name=module_name,
                report_name=report_name,
                button_label=f"📥 {report_name}",
                key=f"dl_{module_name}_{report_name}_{idx}"
            )

