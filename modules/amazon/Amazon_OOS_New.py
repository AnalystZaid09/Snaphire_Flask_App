import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill, Font

from common.ui_utils import (
    apply_professional_style, 
    get_download_filename, 
    render_header,
    download_module_report,
    to_excel
)

MODULE_NAME = "amazon"

# Page configuration
st.set_page_config(
    page_title="Amazon OOS Inventory",
    page_icon="📦",
    layout="wide"
)
apply_professional_style()


# Helper Functions
def normalize_sku(series):
    return series.astype(str)

def add_grand_total_row(df, numeric_cols=None):
    """Add a single Grand Total row with label only in first column"""
    if df.empty:
        return df

    total_row = {}

    # First column name
    first_col = df.columns[0]

    for col in df.columns:
        if col == first_col:
            # Put Grand Total only in first column
            total_row[col] = "Grand Total"

        elif numeric_cols is None or col in numeric_cols:
            # Sum numeric columns
            col_numeric = pd.to_numeric(df[col], errors="coerce")
            if col_numeric.notna().any():
                total_val = col_numeric.sum()
                total_row[col] = round(total_val, 2) if isinstance(total_val, float) else total_val
            else:
                total_row[col] = ""

        else:
            # Keep all other columns blank
            total_row[col] = ""

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

def create_stock_pivot(df):
    """
    Creates pivot table with:
    Rows   : Brand, SKU
    Values : Sum of DOC, DRR, CP
    Adds Brand subtotal + Grand Total
    """
    if df.empty:
        return pd.DataFrame()

    pivot = pd.pivot_table(
        df,
        index=["Brand", "(Parent) ASIN", "SKU"],
        values=["DOC", "DRR", "CP"],
        aggfunc="sum"
    )

    # Reset index for flat table
    pivot = pivot.reset_index()

    # Rename columns for clarity
    pivot.rename(
        columns={
            "DOC": "Sum of DOC",
            "DRR": "Sum of DRR",
            "CP": "Sum of CP"
        },
        inplace=True
    )
    return pivot


def create_inventory_pivot(df):
    """
    Creates pivot table with:
    Rows   : Brand, SKU
    Values : Sum of DOC, DRR, CP
    Adds Brand subtotal + Grand Total
    """
    if df.empty:
        return pd.DataFrame()

    pivot = pd.pivot_table(
        df,
        index=["Brand", "asin", "sku"],
        values=["DOC", "DRR", "CP"],
        aggfunc="sum"
    )

    # Reset index for flat table
    pivot = pivot.reset_index()

    # Rename columns for clarity
    pivot.rename(
        columns={
            "DOC": "Sum of DOC",
            "DRR": "Sum of DRR",
            "CP": "Sum of CP"
        },
        inplace=True
    )
    return pivot

def get_doc_color_hex(doc_value):
    """Return hex color based on DOC value"""
    if pd.isna(doc_value) or doc_value == 0 or doc_value == '':
        return None
    
    try:
        doc_val = float(doc_value)
    except:
        return None
    
    if doc_val < 7:
        return 'FF4444'  # Red
    elif doc_val < 15:
        return 'FF8800'  # Orange
    elif doc_val < 30:
        return '44FF44'  # Green
    elif doc_val < 45:
        return 'FFFF44'  # Yellow
    elif doc_val < 60:
        return '44DDFF'  # Sky Blue
    elif doc_val < 90:
        return '8B4513'  # Brown
    else:
        return '000000'  # Black

def get_doc_font_color(doc_value):
    """Return font color (black or white) based on DOC value for readability"""
    if pd.isna(doc_value) or doc_value == 0 or doc_value == '':
        return '000000'
    
    try:
        doc_val = float(doc_value)
    except:
        return '000000'
    
    # White text for darker backgrounds
    if doc_val < 7 or doc_val < 15 or doc_val >= 90 or (doc_val >= 60 and doc_val < 90):
        return 'FFFFFF'
    else:
        return '000000'

def get_doc_color(doc_value):
    """Return color based on DOC value for Streamlit display"""
    if pd.isna(doc_value) or doc_value == 0 or doc_value == '':
        return ''
    
    try:
        doc_val = float(doc_value)
    except:
        return ''
    
    if doc_val < 7:
        return 'background-color: #ff4444; color: white;'  # Red
    elif doc_val < 15:
        return 'background-color: #ff8800; color: white;'  # Orange
    elif doc_val < 30:
        return 'background-color: #44ff44; color: black;'  # Green
    elif doc_val < 45:
        return 'background-color: #ffff44; color: black;'  # Yellow
    elif doc_val < 60:
        return 'background-color: #44ddff; color: black;'  # Sky Blue
    elif doc_val < 90:
        return 'background-color: #8b4513; color: white;'  # Brown
    else:
        return 'background-color: #000000; color: white;'  # Black

def style_doc_column(df):
    """Apply color styling to DOC column for Streamlit display"""
    def apply_color(val):
        return get_doc_color(val)
    
    # Create a copy to avoid modifying original
    styled = df.style.applymap(apply_color, subset=['DOC'])
    return styled

def to_excel(df, apply_doc_formatting=False):
    """Convert dataframe to Excel bytes with optional DOC color formatting"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        if apply_doc_formatting and 'DOC' in df.columns:
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Find DOC column index
            doc_col_idx = df.columns.get_loc('DOC') + 1  # +1 because Excel is 1-indexed
            
            # Apply formatting to each cell in DOC column (skip header)
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                cell = worksheet.cell(row=row_idx, column=doc_col_idx)
                doc_value = df.iloc[row_idx - 2]['DOC']
                
                bg_color = get_doc_color_hex(doc_value)
                font_color = get_doc_font_color(doc_value)
                
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    cell.font = Font(color=font_color, bold=True)
    
    return output.getvalue()

def process_business_report(business_file, purchase_master_file, inventory_file, listing_file, no_of_days, doc_threshold):
    """Process all business report data"""
    
    # Read Business Report
    if business_file.name.endswith('.csv'):
        Business_Report = pd.read_csv(business_file)
    else:
        Business_Report = pd.read_excel(business_file)
    Business_Report["SKU"] = normalize_sku(Business_Report["SKU"])
    
    # Clean Total Order Items
    Business_Report["Total Order Items"] = (
        Business_Report["Total Order Items"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    Business_Report["Total Order Items"] = pd.to_numeric(
        Business_Report["Total Order Items"], errors="coerce"
    )
    
    # Calculate Total Sales Order
    Business_Report["Total Sales Order"] = (
        Business_Report["Total Order Items"] + 
        Business_Report["Total Order Items - B2B"]
    )
    
    # Create Business Pivot
    Business_Pivot = pd.pivot_table(
        Business_Report,
        index=["SKU","(Parent) ASIN"],
        values="Total Sales Order",
        aggfunc="sum"
    )
    Business_Pivot = Business_Pivot.reset_index()
    
    # Sort by Total Sales Order
    Business_Pivot = Business_Pivot.sort_values("Total Sales Order", ascending=False)
    
    # Read Purchase Master
    if purchase_master_file.name.endswith('.csv'):
        purchase_master = pd.read_csv(purchase_master_file)
    else:
        purchase_master = pd.read_excel(purchase_master_file)
    purchase_master["Amazon Sku Name"] = normalize_sku(purchase_master["Amazon Sku Name"])
    
    # Map Purchase Master data
    purchase_master1 = purchase_master.iloc[:, [2, 3, 4, 6, 7]].copy()
    purchase_master1.columns = [
        "Amazon Sku Name", "Vendor SKU Codes", "Brand Manager", "Brand", "Product Name"
    ]
    purchase_master1 = purchase_master1.drop_duplicates(subset="Amazon Sku Name", keep="first")
    
    Business_Pivot["(Parent) ASIN"] = Business_Pivot["(Parent) ASIN"].astype(str).str.strip()
    purchase_master["ASIN"] = purchase_master["ASIN"].astype(str).str.strip()

    # Step-by-step mapping from Purchase Master using ASIN
    # Create lookup dictionaries based on ASIN
    pm_asin_lookup = purchase_master.drop_duplicates(subset="ASIN", keep="first")
    
    vendor_sku_map = pm_asin_lookup.set_index("ASIN")["Vendor SKU Codes"]
    brand_map = pm_asin_lookup.set_index("ASIN")["Brand"]
    product_map = pm_asin_lookup.set_index("ASIN")["Product Name"]
    manager_map = pm_asin_lookup.set_index("ASIN")["Brand Manager"]
    
    Business_Pivot["Vendor SKU Codes"] = Business_Pivot["(Parent) ASIN"].map(vendor_sku_map)
    Business_Pivot["Brand"] = Business_Pivot["(Parent) ASIN"].map(brand_map)
    Business_Pivot["Product Name"] = Business_Pivot["(Parent) ASIN"].map(product_map)
    Business_Pivot["Brand Manager"] = Business_Pivot["(Parent) ASIN"].map(manager_map)
    
    # Aggregate CP by ASIN (Ensure numeric first)
    purchase_master["CP"] = pd.to_numeric(purchase_master["CP"], errors="coerce").fillna(0)
    purchase_master_cp = (
        purchase_master.copy()
        .groupby("ASIN", as_index=False)["CP"]
        .sum()
    )

    # Set index for mapping
    purchase_master_cp = purchase_master_cp.set_index("ASIN")

    # Map summed CP to pivot
    Business_Pivot["CP"] = Business_Pivot["(Parent) ASIN"].map(purchase_master_cp["CP"])

    # Fill NaN in CP with 0, ensure numeric, and round to 2 decimals
    Business_Pivot["CP"] = pd.to_numeric(Business_Pivot["CP"], errors="coerce").fillna(0).round(2)
    
    # Reorder columns
    Business_Pivot = Business_Pivot[[
        "SKU","(Parent) ASIN", "Vendor SKU Codes", "Brand", "Product Name", 
        "Brand Manager", "Total Sales Order", "CP"
    ]]
    
    Business_Pivot.fillna("", inplace=True)
    
    # Calculate As Per Qty
    Business_Pivot["As Per Qty"] = Business_Pivot["Total Sales Order"] * Business_Pivot["CP"]
    
    # Calculate DRR and round to 2 decimal places
    Business_Pivot["DRR"] = (Business_Pivot["Total Sales Order"] / no_of_days).round(2)
    
    # Read Inventory
    if inventory_file.name.endswith('.csv'):
        Inventory = pd.read_csv(inventory_file)
    else:
        Inventory = pd.read_excel(inventory_file)
    Inventory["sku"] = normalize_sku(Inventory["sku"])
    
    # Create Inventory Pivot
    inventory_pivot = pd.pivot_table(
        Inventory,
        index="asin",
        values=["afn-fulfillable-quantity", "afn-reserved-quantity"],
        aggfunc="sum"
    )
    inventory_pivot = inventory_pivot.reset_index()
    inventory_pivot.rename(
        columns={
            "afn-fulfillable-quantity": "afn-fulfillable-qty",
            "afn-reserved-quantity": "afn-reserved-qty"
        },
        inplace=True
    )
    inventory_pivot["Total Stock"] = (
        inventory_pivot["afn-fulfillable-qty"] + 
        inventory_pivot["afn-reserved-qty"]
    )
    
    # Map inventory to Business Pivot
    inventory_lookup = inventory_pivot.drop_duplicates(subset="asin", keep="first").set_index("asin")
    Business_Pivot["afn-fulfillable-qty"] = Business_Pivot["(Parent) ASIN"].map(inventory_lookup["afn-fulfillable-qty"])
    Business_Pivot["afn-reserved-qty"] = Business_Pivot["(Parent) ASIN"].map(inventory_lookup["afn-reserved-qty"])
    Business_Pivot["Total Stock"] = Business_Pivot["(Parent) ASIN"].map(inventory_lookup["Total Stock"])
    
    # Calculate DOC
    Business_Pivot["DOC"] = Business_Pivot["Total Stock"] / Business_Pivot["DRR"]
    Business_Pivot["DOC"] = Business_Pivot["DOC"].replace([np.inf, -np.inf], np.nan)
    Business_Pivot["DOC"] = Business_Pivot["DOC"].apply(
        lambda x: round(x, 2) if pd.notna(x) else ""
    )
    
    # Process Listing Report
    if listing_file is not None:
        if listing_file.name.endswith('.csv'):
            Listing_Status = pd.read_csv(listing_file)
        else:
            Listing_Status = pd.read_excel(listing_file)
        
        seller_sku_series = Listing_Status.iloc[:, 3].astype(str)
        seller_sku_lookup = dict(zip(seller_sku_series, seller_sku_series))
        Business_Pivot["seller-sku"] = Business_Pivot["SKU"].map(seller_sku_lookup)
        Business_Pivot["Closing Listing"] = (
            Business_Pivot["seller-sku"]
            .eq(Business_Pivot["SKU"])
            .map({True: "Closing", False: ""})
        )
    
    # Add Grand Total to Business Pivot
    numeric_cols = ["Total Sales Order", "CP", "As Per Qty", "DRR", "afn-fulfillable-qty", "afn-reserved-qty", "Total Stock"]
    Business_Pivot = add_grand_total_row(Business_Pivot, numeric_cols)
    
    # Create OOS Report (before adding grand total)
    OOS_Report = Business_Pivot[Business_Pivot["SKU"] != "Grand Total"].copy()
    OOS_Report = OOS_Report[OOS_Report["afn-fulfillable-qty"] == 0].reset_index(drop=True)
    
    # Ensure numeric columns for pivot and round to 2 decimals
    OOS_Report["CP"] = pd.to_numeric(OOS_Report["CP"], errors="coerce").fillna(0).round(2)
    OOS_Report["DOC"] = pd.to_numeric(OOS_Report["DOC"], errors="coerce").fillna(0).round(2)
    OOS_Report["DRR"] = pd.to_numeric(OOS_Report["DRR"], errors="coerce").fillna(0).round(2)
    
    # Create Overstock Report (before adding grand total)
    Overstock_Report = Business_Pivot[Business_Pivot["SKU"] != "Grand Total"].copy()
    Overstock_Report["DOC_compare"] = pd.to_numeric(Overstock_Report["DOC"], errors="coerce").fillna(0).round(2)
    Overstock_Report["CP"] = pd.to_numeric(Overstock_Report["CP"], errors="coerce").fillna(0).round(2)
    Overstock_Report["DRR"] = pd.to_numeric(Overstock_Report["DRR"], errors="coerce").fillna(0).round(2)
    Overstock_Report = Overstock_Report[Overstock_Report["DOC_compare"] > doc_threshold].reset_index(drop=True)
    Overstock_Report = Overstock_Report.drop("DOC_compare", axis=1)
    
    # Add grand totals to reports (include DOC)
    numeric_cols = ["Total Sales Order", "CP", "As Per Qty", "DRR", "afn-fulfillable-qty", "afn-reserved-qty", "Total Stock", "DOC"]
    OOS_Report = add_grand_total_row(OOS_Report, numeric_cols)
    Overstock_Report = add_grand_total_row(Overstock_Report, numeric_cols)
    
    # Create pivots for OOS and Overstock (they already have margins=True for grand total)
    # OOS_Pivot = create_stock_pivot(OOS_Report[OOS_Report["SKU"] != "Grand Total"])
    # Overstock_Pivot = create_stock_pivot(Overstock_Report[Overstock_Report["SKU"] != "Grand Total"])
    # --- Business OOS Pivot ---
    oos_pivot_source = OOS_Report[OOS_Report["SKU"] != "Grand Total"].copy()

    OOS_Pivot = create_stock_pivot(oos_pivot_source)

    OOS_Pivot = add_grand_total_row(
        OOS_Pivot,
        numeric_cols=["Sum of CP", "Sum of DOC", "Sum of DRR"]
    )

    # --- Business Overstock Pivot ---
    overstock_pivot_source = Overstock_Report[Overstock_Report["SKU"] != "Grand Total"].copy()

    Overstock_Pivot = create_stock_pivot(overstock_pivot_source)

    Overstock_Pivot = add_grand_total_row(
        Overstock_Pivot,
        numeric_cols=["Sum of CP", "Sum of DOC", "Sum of DRR"]
    )

    
    return Business_Pivot, OOS_Report, Overstock_Report, OOS_Pivot, Overstock_Pivot

def process_inventory_report(inventory_file, purchase_master_file, business_pivot, no_of_days_inventory, doc_inventory_threshold):
    """Process inventory report data"""
    
    # Read Inventory
    if inventory_file.name.endswith('.csv'):
        Inventory = pd.read_csv(inventory_file)
    else:
        Inventory = pd.read_excel(inventory_file)
    Inventory["sku"] = normalize_sku(Inventory["sku"])
    
    # First aggregate by ASIN to get total quantities per ASIN
    inventory_asin_totals = Inventory.groupby("asin", as_index=False).agg({
        "afn-fulfillable-quantity": "sum",
        "afn-reserved-quantity": "sum"
    })
    
    # Get unique SKUs per ASIN (take first SKU for each ASIN)
    inventory_sku_map = Inventory.groupby("asin", as_index=False)["sku"].first()
    
    # Merge to create Inventory Report Pivot
    Inventory_Report_Pivot = inventory_asin_totals.merge(inventory_sku_map, on="asin", how="left")
    
    Inventory_Report_Pivot["Total Stock"] = (
        Inventory_Report_Pivot["afn-fulfillable-quantity"] + 
        Inventory_Report_Pivot["afn-reserved-quantity"]
    )
    
    # Read Purchase Master
    if purchase_master_file.name.endswith('.csv'):
        purchase_master = pd.read_csv(purchase_master_file)
    else:
        purchase_master = pd.read_excel(purchase_master_file)
    purchase_master["Amazon Sku Name"] = normalize_sku(purchase_master["Amazon Sku Name"])
    
    # Map Purchase Master data
    pm_lookup = purchase_master[[
        "ASIN","Amazon Sku Name", "Vendor SKU Codes", "Brand", 
        "Brand Manager", "Product Name", "CP"
    ]].copy()
    pm_lookup = pm_lookup.drop_duplicates(subset="ASIN", keep="first").set_index("ASIN")
    
    Inventory_Report_Pivot["Vendor SKU Codes"] = Inventory_Report_Pivot["asin"].map(pm_lookup["Vendor SKU Codes"])
    Inventory_Report_Pivot["Brand"] = Inventory_Report_Pivot["asin"].map(pm_lookup["Brand"])
    Inventory_Report_Pivot["Brand Manager"] = Inventory_Report_Pivot["asin"].map(pm_lookup["Brand Manager"])
    Inventory_Report_Pivot["Product Name"] = Inventory_Report_Pivot["asin"].map(pm_lookup["Product Name"])
    Inventory_Report_Pivot["CP"] = Inventory_Report_Pivot["asin"].map(pm_lookup["CP"])
    
    # Reorder columns
    Inventory_Report_Pivot = Inventory_Report_Pivot[[
        "asin","sku", "Vendor SKU Codes", "Brand", "Brand Manager", 
        "Product Name", "afn-fulfillable-quantity", 
        "afn-reserved-quantity", "Total Stock", "CP"
    ]]
    
    # Normalize both ASIN columns to string
    business_pivot["(Parent) ASIN"] = business_pivot["(Parent) ASIN"].astype(str).str.strip()
    Inventory_Report_Pivot["asin"] = Inventory_Report_Pivot["asin"].astype(str).str.strip()

    # Create lookup with unique index by summing duplicates
    business_lookup = business_pivot[business_pivot["SKU"] != "Grand Total"].groupby("(Parent) ASIN", as_index=False)["Total Sales Order"].sum()
    business_lookup = business_lookup.set_index("(Parent) ASIN")
    business_lookup.index = business_lookup.index.astype(str)

    # Map using ASIN instead of SKU
    Inventory_Report_Pivot["Total Sales Order"] = Inventory_Report_Pivot["asin"].map(
        business_lookup["Total Sales Order"]
    )
    
    # Fill NaN values
    Inventory_Report_Pivot[["CP", "Total Sales Order"]] = (
        Inventory_Report_Pivot[["CP", "Total Sales Order"]].fillna(0)
    )
    
    # Clean CP
    Inventory_Report_Pivot["CP"] = (
        Inventory_Report_Pivot["CP"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    Inventory_Report_Pivot["CP"] = pd.to_numeric(
        Inventory_Report_Pivot["CP"], errors="coerce"
    ).round(2)
    
    # Calculate As Per Qty
    Inventory_Report_Pivot["As Per Qty"] = (
        Inventory_Report_Pivot["CP"] * Inventory_Report_Pivot["Total Sales Order"]
    ).round(2)
    
    # Calculate DRR
    Inventory_Report_Pivot["DRR"] = (
        Inventory_Report_Pivot["Total Sales Order"] / no_of_days_inventory
    ).round(2)
    
    # Calculate DOC
    Inventory_Report_Pivot["DOC"] = (
        Inventory_Report_Pivot["Total Stock"] / Inventory_Report_Pivot["DRR"]
    )
    Inventory_Report_Pivot["DOC"] = Inventory_Report_Pivot["DOC"].replace(
        [np.inf, -np.inf], np.nan
    )
    # Keep DOC as numeric but handle NaN
    Inventory_Report_Pivot["DOC"] = Inventory_Report_Pivot["DOC"].apply(
        lambda x: round(x, 2) if pd.notna(x) else 0
    )
    
    # DON'T filter out items with no sales - show all inventory
    # Inventory_Report_Pivot = Inventory_Report_Pivot[
    #     Inventory_Report_Pivot["Total Sales Order"] != 0
    # ]
    
    # Create OOS Inventory (before adding grand total)
    OOS_Inventory = Inventory_Report_Pivot[
        Inventory_Report_Pivot["afn-fulfillable-quantity"] == 0
    ].copy().reset_index(drop=True)
    
    # Create Overstock Inventory (before adding grand total)
    # Ensure DOC is numeric for comparison
    Overstock_Inventory = Inventory_Report_Pivot[
        Inventory_Report_Pivot["DOC"] >= doc_inventory_threshold
    ].copy().reset_index(drop=True)
    
    # Add Grand Total to all reports (include DOC in numeric columns)
    numeric_cols = ["afn-fulfillable-quantity", "afn-reserved-quantity", "Total Stock", "CP", "Total Sales Order", "As Per Qty", "DRR", "DOC"]
    Inventory_Report_Pivot = add_grand_total_row(Inventory_Report_Pivot, numeric_cols)
    OOS_Inventory = add_grand_total_row(OOS_Inventory, numeric_cols)
    Overstock_Inventory = add_grand_total_row(Overstock_Inventory, numeric_cols)
    
    # # Add grand total to Overstock Inventory
    # Overstock_Inventory = add_grand_total_row(Overstock_Inventory, numeric_cols)
    
    # Create pivots (they already have margins=True for grand total)
    # OOS_Inventory_Pivot = create_inventory_pivot(OOS_Inventory[OOS_Inventory["asin"] != "Grand Total"])
    # Overstock_Inventory_Pivot = create_inventory_pivot(Overstock_Inventory[Overstock_Inventory["asin"] != "Grand Total"])
    
    # -------- Inventory OOS Pivot --------
    oos_inventory_pivot_source = OOS_Inventory[
        (OOS_Inventory["asin"] != "Grand Total") &
        (OOS_Inventory["sku"] != "Grand Total")
    ].copy()

    OOS_Inventory_Pivot = create_inventory_pivot(oos_inventory_pivot_source)

    OOS_Inventory_Pivot = add_grand_total_row(
        OOS_Inventory_Pivot,
        numeric_cols=["Sum of CP", "Sum of DOC", "Sum of DRR"]
    )

    # -------- Inventory Overstock Pivot --------
    overstock_inventory_pivot_source = Overstock_Inventory[
        (Overstock_Inventory["asin"] != "Grand Total") &
        (Overstock_Inventory["sku"] != "Grand Total")
    ].copy()

    Overstock_Inventory_Pivot = create_inventory_pivot(overstock_inventory_pivot_source)

    Overstock_Inventory_Pivot = add_grand_total_row(
        Overstock_Inventory_Pivot,
        numeric_cols=["Sum of CP", "Sum of DOC", "Sum of DRR"]
    )

    return Inventory_Report_Pivot, OOS_Inventory, Overstock_Inventory, OOS_Inventory_Pivot, Overstock_Inventory_Pivot

# Main App
# Main App
render_header("Amazon OOS Inventory Management System", None)

# Sidebar for file uploads
st.sidebar.header("📁 Upload Files")

business_file = st.sidebar.file_uploader("Business Report", type=['csv', 'xlsx'])
purchase_master_file = st.sidebar.file_uploader("Purchase Master", type=['csv', 'xlsx'])
inventory_file = st.sidebar.file_uploader("Manage Inventory", type=['csv', 'xlsx'])
listing_file = st.sidebar.file_uploader("Listing Status (Optional)", type=['csv', 'xlsx'])

st.sidebar.header("⚙️ Parameters")
no_of_days = st.sidebar.number_input("Number of Days (Business)", min_value=1, value=30)
doc_threshold = st.sidebar.number_input("DOC Threshold (Business)", min_value=0, value=30)
no_of_days_inventory = st.sidebar.number_input("Number of Days (Inventory)", min_value=1, value=30)
doc_inventory_threshold = st.sidebar.number_input("DOC Threshold (Inventory)", min_value=0, value=30)

# DOC Legend
st.sidebar.markdown("---")
st.sidebar.markdown("### 🎨 DOC Color Legend")
st.sidebar.markdown("🔴 **Red (0-7)**: Critical")
st.sidebar.markdown("🟠 **Orange (7-15)**: Low")
st.sidebar.markdown("🟢 **Green (15-30)**: Optimal")
st.sidebar.markdown("🟡 **Yellow (30-45)**: Monitor")
st.sidebar.markdown("🔵 **Sky Blue (45-60)**: High")
st.sidebar.markdown("🟤 **Brown (60-90)**: Excess")
st.sidebar.markdown("⬛ **Black (>90)**: Overstocked")

# Process data when all required files are uploaded
if business_file and purchase_master_file and inventory_file:
    try:
        # Create copies of file objects to avoid pointer issues
        import io
        
        # Reset all file pointers to beginning
        business_file.seek(0)
        purchase_master_file.seek(0)
        inventory_file.seek(0)
        if listing_file:
            listing_file.seek(0)
        
        # Read files into memory to avoid pointer issues
        business_bytes = business_file.read()
        purchase_bytes = purchase_master_file.read()
        inventory_bytes = inventory_file.read()
        listing_bytes = listing_file.read() if listing_file else None
        
        # Create new file-like objects
        business_io = io.BytesIO(business_bytes)
        purchase_io = io.BytesIO(purchase_bytes)
        inventory_io = io.BytesIO(inventory_bytes)
        listing_io = io.BytesIO(listing_bytes) if listing_bytes else None
        
        # Set names for the file objects
        business_io.name = business_file.name
        purchase_io.name = purchase_master_file.name
        inventory_io.name = inventory_file.name
        if listing_io:
            listing_io.name = listing_file.name
        
        # Process Business Report
        Business_Pivot, OOS_Report, Overstock_Report, OOS_Pivot, Overstock_Pivot = process_business_report(
            business_io, purchase_io, inventory_io, listing_io, 
            no_of_days, doc_threshold
        )
        
        # Reset file pointers for inventory processing
        purchase_io.seek(0)
        inventory_io.seek(0)
        
        # Process Inventory Report
        Inventory_Report_Pivot, OOS_Inventory, Overstock_Inventory, OOS_Inventory_Pivot, Overstock_Inventory_Pivot = process_inventory_report(
            inventory_io, purchase_io, Business_Pivot, 
            no_of_days_inventory, doc_inventory_threshold
        )
        
        # Create tabs
        tab1, tab2, tab3 = st.tabs(["📊 Business Report", "📦 Inventory Report", "📋 Business Listing Report"])
        
        # Tab 1: Business Report
        with tab1:
            st.header("Business Report")
            
            # Sub-tabs for Business Report
            sub_tab1, sub_tab2, sub_tab3 = st.tabs(["Main Report", "OOS Report", "Overstock Report"])
            
            with sub_tab1:
                st.subheader("Business Pivot Report")
                
                # Apply DOC color formatting
                if 'DOC' in Business_Pivot.columns:
                    styled_df = style_doc_column(Business_Pivot)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Business_Pivot, use_container_width=True, height=600)
                
                st.download_button(
                    "📥 Download Business Pivot (with DOC colors)",
                    data=to_excel(Business_Pivot, apply_doc_formatting=True),
                    file_name=get_download_filename("business_pivot"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Save to MongoDB (Best effort)
                try:
                    with st.spinner("Saving Business Pivot to Database..."):
                        save_reconciliation_report(
                            collection_name="amazon_oos_business",
                            invoice_no=f"OOS_Report_{datetime.now().strftime('%Y%m%d%H%M')}",
                            summary_data=pd.DataFrame(), # No specific summary df, passing empty
                            line_items_data=Business_Pivot, # Saving the main pivot
                            metadata={"type": "business_pivot"}
                        )
                except Exception as e:
                    # Don't block user flow
                    print(f"DB Save error: {e}")
            
            with sub_tab2:
                st.subheader("Out of Stock (OOS) Report")
                
                # Apply DOC color formatting
                if 'DOC' in OOS_Report.columns:
                    styled_df = style_doc_column(OOS_Report)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(OOS_Report, use_container_width=True, height=600)
                
                st.download_button(
                    "📥 Download OOS Report (with DOC colors)",
                    data=to_excel(OOS_Report, apply_doc_formatting=True),
                    file_name=get_download_filename("oos_report"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.subheader("OOS Pivot Table")
                st.dataframe(OOS_Pivot, use_container_width=True)
                download_module_report(
                    df=OOS_Pivot,
                    module_name=MODULE_NAME,
                    report_name="OOS Pivot",
                    button_label="📥 Download OOS Pivot",
                    key="oos_pivot_download"
                )
            
            with sub_tab3:
                st.subheader("Overstock Report")
                
                # Apply DOC color formatting
                if 'DOC' in Overstock_Report.columns:
                    styled_df = style_doc_column(Overstock_Report)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Overstock_Report, use_container_width=True, height=600)
                
                st.download_button(
                    "📥 Download Overstock Report (with DOC colors)",
                    data=to_excel(Overstock_Report, apply_doc_formatting=True),
                    file_name=get_download_filename("overstock_report"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.subheader("Overstock Pivot Table")
                st.dataframe(Overstock_Pivot, use_container_width=True)
                st.download_button(
                    "📥 Download Overstock Pivot",
                    data=to_excel(Overstock_Pivot),
                    file_name=get_download_filename("overstock_pivot"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # Tab 2: Inventory Report
        with tab2:
            st.header("Inventory Report")
            
            # Sub-tabs for Inventory Report
            sub_tab1, sub_tab2, sub_tab3 = st.tabs(["Main Report", "OOS Inventory", "Overstock Inventory"])
            
            with sub_tab1:
                st.subheader("Inventory Report Pivot")
                
                # Apply DOC color formatting
                if 'DOC' in Inventory_Report_Pivot.columns:
                    styled_df = style_doc_column(Inventory_Report_Pivot)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Inventory_Report_Pivot, use_container_width=True, height=600)
                
                st.download_button(
                    "📥 Download Inventory Report (with DOC colors)",
                    data=to_excel(Inventory_Report_Pivot, apply_doc_formatting=True),
                    file_name=get_download_filename("inventory_report"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Save to MongoDB (Best effort)
                try:
                     save_reconciliation_report(
                        collection_name="amazon_oos_inventory",
                        invoice_no=f"Inventory_Report_{datetime.now().strftime('%Y%m%d%H%M')}",
                        summary_data=pd.DataFrame(),
                        line_items_data=Inventory_Report_Pivot,
                        metadata={"type": "inventory_pivot"}
                    )
                except Exception as e:
                    pass
            
            with sub_tab2:
                st.subheader("OOS Inventory Report")
                
                # Apply DOC color formatting
                if 'DOC' in OOS_Inventory.columns:
                    styled_df = style_doc_column(OOS_Inventory)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(OOS_Inventory, use_container_width=True, height=600)
                
                download_module_report(
                    df=OOS_Inventory,
                    module_name=MODULE_NAME,
                    report_name="OOS Inventory (with DOC colors)",
                    button_label="📥 Download OOS Inventory (with DOC colors)",
                    apply_doc_formatting=True,
                    key="oos_inventory_download"
                )
                
                st.subheader("OOS Inventory Pivot Table")
                st.dataframe(OOS_Inventory_Pivot, use_container_width=True)
                download_module_report(
                    df=OOS_Inventory_Pivot,
                    module_name=MODULE_NAME,
                    report_name="OOS Inventory Pivot",
                    button_label="📥 Download OOS Inventory Pivot",
                    key="oos_inventory_pivot_download"
                )
            
            with sub_tab3:
                st.subheader("Overstock Inventory Report")
                
                # Apply DOC color formatting
                if 'DOC' in Overstock_Inventory.columns:
                    styled_df = style_doc_column(Overstock_Inventory)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Overstock_Inventory, use_container_width=True, height=600)
                
                download_module_report(
                    df=Overstock_Inventory,
                    module_name=MODULE_NAME,
                    report_name="Overstock Inventory (with DOC colors)",
                    button_label="📥 Download Overstock Inventory (with DOC colors)",
                    apply_doc_formatting=True,
                    key="overstock_inventory_download"
                )
                
                st.subheader("Overstock Inventory Pivot Table")
                st.dataframe(Overstock_Inventory_Pivot, use_container_width=True)
                download_module_report(
                    df=Overstock_Inventory_Pivot,
                    module_name=MODULE_NAME,
                    report_name="Overstock Inventory Pivot",
                    button_label="📥 Download Overstock Inventory Pivot",
                    key="overstock_inventory_pivot_download"
                )
        
        # Tab 3: Business Listing Report with DOC coloring
        with tab3:
            st.header("Business Listing Report with DOC Color Coding")
            
            # Apply DOC coloring
            if 'DOC' in Business_Pivot.columns:
                styled_df = style_doc_column(Business_Pivot)
                st.dataframe(styled_df, use_container_width=True, height=600)
            else:
                st.dataframe(Business_Pivot, use_container_width=True, height=600)
            
            download_module_report(
                df=Business_Pivot,
                module_name=MODULE_NAME,
                report_name="Business Listing Report (with DOC colors)",
                button_label="📥 Download Business Listing Report (with DOC colors)",
                apply_doc_formatting=True,
                key="business_listing_download"
            )
            
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        st.exception(e)
else:
    st.info("👆 Please upload all required files (Business Report, Purchase Master, and Manage Inventory) to begin.")
