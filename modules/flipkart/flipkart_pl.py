import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import base64
import gc
from common.mongo import save_reconciliation_report
from datetime import datetime
from common.ui_utils import apply_professional_style, render_header

# Page configuration is handled by the main app
apply_professional_style()

# Initialize session state
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'brand_pivot' not in st.session_state:
    st.session_state.brand_pivot = None
if 'files_hash' not in st.session_state:
    st.session_state.files_hash = None

render_header("Flipkart P&L Analysis", "Upload payment files and product masters to analyze profits")
st.markdown("---")

# File upload section
st.sidebar.header("📁 Upload Files")

payment_file = st.sidebar.file_uploader(
    "Payment File Flipkart",
    type=["xlsx", "csv"],
    help="Upload the Flipkart payment file with Orders sheet (Excel or CSV)"
)

uni_orders_file = st.sidebar.file_uploader(
    "Uni Orders",
    type=["xlsx", "csv"],
    help="Upload the unique orders file (Excel or CSV)"
)

fsn_file = st.sidebar.file_uploader(
    "FSN",
    type=["xlsx", "csv"],
    help="Upload the FSN master file (Excel or CSV)"
)

pm_file = st.sidebar.file_uploader(
    "flipkart PM",
    type=["xlsx", "csv"],
    help="Upload the Product Master file (Excel or CSV)"
)


def get_files_hash(files):
    """Generate a simple hash from file names and sizes to detect changes"""
    hash_str = ""
    for f in files:
        if f is not None:
            hash_str += f"{f.name}_{f.size}_"
    return hash_str if hash_str else None


def get_download_link(df, filename, file_type="excel"):
    """Generate a download link for dataframe - works on Streamlit Cloud"""
    try:
        if file_type == "excel":
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # Apply color formatting to Excel
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                from openpyxl.styles import PatternFill
                
                # Define fill colors (darker shades)
                green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Dark green
                red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  # Dark red
                
                # Columns to format
                profit_columns = [
                    "Profit",
                    "Profit In Percentage",
                    "Profit With Support",
                    "Profit In Percentage With Support",
                    "After 3% Profit",
                    "After 3%"
                ]
                
                # Get column indices
                header_row = list(df.columns)
                col_indices = []
                for col in profit_columns:
                    if col in header_row:
                        col_indices.append(header_row.index(col) + 1)  # openpyxl is 1-indexed
                
                # Apply colors to cells
                for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                    for col_idx in col_indices:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        try:
                            val = pd.to_numeric(cell.value, errors='coerce')
                            if pd.notna(val):
                                if val < 0:
                                    cell.fill = red_fill
                                elif val > 0:
                                    cell.fill = green_fill
                        except:
                            pass
            
            b64 = base64.b64encode(output.getvalue()).decode()
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return f'<a href="data:{mime};base64,{b64}" download="{filename}" class="download-link">📥 Download {filename}</a>'
        else:
            csv = df.to_csv(index=False)
            b64 = base64.b64encode(csv.encode()).decode()
            mime = "text/csv"
            return f'<a href="data:{mime};base64,{b64}" download="{filename}" class="download-link">📄 Download {filename}</a>'
    except Exception as e:
        return f"Error generating download: {str(e)}"


def color_profit_cells(val):
    """Apply background color to cells based on value: green for positive, red for negative"""
    try:
        num_val = pd.to_numeric(val, errors='coerce')
        if pd.isna(num_val) or val == "" or val == "Grand Total":
            return ''
        elif num_val < 0:
            return 'background-color: #FF6B6B'  # Dark red
        elif num_val > 0:
            return 'background-color: #92D050'  # Dark green
        else:
            return ''
    except:
        return ''


def style_dataframe(df):
    """Apply styling to the dataframe with color-coded profit columns"""
    # Columns to apply color formatting
    profit_columns = [
        "Profit",
        "Profit In Percentage",
        "Profit With Support",
        "Profit In Percentage With Support",
        "After 3% Profit",
        "After 3%"
    ]
    
    # Only style columns that exist in the dataframe
    cols_to_style = [col for col in profit_columns if col in df.columns]
    
    if cols_to_style:
        return df.style.applymap(color_profit_cells, subset=cols_to_style)
    return df.style


def process_data(payment_file, uni_orders_file, fsn_file, pm_file):
    """Process all uploaded files and generate P&L report"""
    
    try:
        # 1. Read Payment File with custom header
        if payment_file.name.endswith('.csv'):
            payment_df = pd.read_csv(payment_file, header=None)
        else:
            payment_df = pd.read_excel(payment_file, sheet_name="Orders", header=None)
        
        payment_df.columns = payment_df.iloc[1]
        payment_df = payment_df.drop(index=[0, 1, 2]).reset_index(drop=True)
        
        # Clean column names
        payment_df.columns = (
            payment_df.columns
            .astype(str)
            .str.strip()
            .str.replace("\n", " ", regex=False)
            .str.replace("  ", " ", regex=False)
        )
        
        # 2. Read Unique Orders
        if uni_orders_file.name.endswith('.csv'):
            unique_order = pd.read_csv(uni_orders_file)
        else:
            unique_order = pd.read_excel(uni_orders_file)
        
        # Add Unique Order ID column
        payment_df["Unique Order ID"] = payment_df["Order ID"].where(
            payment_df["Order ID"].isin(unique_order["Order ID"]),
            np.nan
        )
        
        # 3. Filter by Protection Fund = 0
        payment_df["Protection Fund (Rs.)"] = (
            payment_df["Protection Fund (Rs.)"]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
        )
        payment_df["Protection Fund (Rs.)"] = pd.to_numeric(
            payment_df["Protection Fund (Rs.)"],
            errors="coerce"
        )
        payment_df = payment_df[
            payment_df["Protection Fund (Rs.)"].notna() &
            (payment_df["Protection Fund (Rs.)"] == 0)
        ].reset_index(drop=True)
        
        # 4. Filter by Refund = 0
        payment_df["Refund (Rs.)"] = (
            payment_df["Refund (Rs.)"]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
        )
        payment_df["Refund (Rs.)"] = pd.to_numeric(
            payment_df["Refund (Rs.)"],
            errors="coerce"
        )
        payment_df = payment_df[
            payment_df["Refund (Rs.)"].notna() &
            (payment_df["Refund (Rs.)"] == 0)
        ].reset_index(drop=True)
        
        # 5. Keep only NEW orders (Unique Order ID is NaN)
        payment_df = payment_df[payment_df["Unique Order ID"].isna()]
        
        # 6. Remove NaN columns and Unique Order ID
        payment_df = payment_df.loc[:, ~(
            payment_df.columns.isna() |
            (payment_df.columns.astype(str).str.lower() == 'nan')
        )]
        payment_df = payment_df.drop(columns=['Unique Order ID'], errors='ignore')
        
        # 7. Create working copy
        payment_df_filter = payment_df.copy()
        
        # Clear original to save memory
        del payment_df
        gc.collect()
        
        # 8. Convert numeric columns
        cols_to_numeric = ["Sale Amount (Rs.)", "Total Offer Amount (Rs.)", "My share (Rs.)"]
        for col in cols_to_numeric:
            if col in payment_df_filter.columns:
                payment_df_filter[col] = (
                    payment_df_filter[col]
                    .astype(str)
                    .str.replace(",", "", regex=False)
                    .str.strip()
                )
                payment_df_filter[col] = pd.to_numeric(payment_df_filter[col], errors="coerce").fillna(0)
        
        # 9. Calculate Sales Amount
        payment_df_filter["Sales Amount"] = (
            payment_df_filter["Sale Amount (Rs.)"]
            + payment_df_filter["Total Offer Amount (Rs.)"]
            + payment_df_filter["My share (Rs.)"]
        )
        
        # Position Sales Amount after My share
        my_share_idx = payment_df_filter.columns.get_loc("My share (Rs.)")
        sales_amount_col = payment_df_filter.pop("Sales Amount")
        payment_df_filter.insert(my_share_idx + 1, "Sales Amount", sales_amount_col)
        
        # 10. Clean and filter Sales Amount
        payment_df_filter["Sales Amount"] = (
            payment_df_filter["Sales Amount"]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
        )
        payment_df_filter["Sales Amount"] = pd.to_numeric(
            payment_df_filter["Sales Amount"],
            errors="coerce"
        ).round(2)
        payment_df_filter = payment_df_filter[payment_df_filter["Sales Amount"] > 0.0]
        
        # 11. Convert Bank Settlement Value
        value_cols = ["Sales Amount", "Bank Settlement Value (Rs.) = SUM(J:R)"]
        for col in value_cols:
            if col in payment_df_filter.columns:
                payment_df_filter[col] = (
                    payment_df_filter[col]
                    .astype(str)
                    .str.replace(",", "", regex=False)
                    .str.strip()
                )
                payment_df_filter[col] = pd.to_numeric(payment_df_filter[col], errors="coerce").fillna(0)
        
        # 12. Create Pivot Table
        pivot_df = pd.pivot_table(
            payment_df_filter,
            index=["Order Date", "Order item ID", "Order ID", "Seller SKU"],
            values=["Sales Amount", "Bank Settlement Value (Rs.) = SUM(J:R)"],
            aggfunc="sum"
        ).reset_index()
        
        # 13. Add Quantity lookup
        pivot_df["Order item ID"] = pivot_df["Order item ID"].astype(str).str.strip()
        payment_df_filter["Order item ID"] = payment_df_filter["Order item ID"].astype(str).str.strip()
        
        if "Quantity" in payment_df_filter.columns:
            lookup_dict = (
                payment_df_filter
                .drop_duplicates("Order item ID")
                .set_index("Order item ID")["Quantity"]
            )
            pivot_df["Sum of Quantity"] = pivot_df["Order item ID"].map(lookup_dict)
        else:
            pivot_df["Sum of Quantity"] = 1
        
        # Clear to save memory
        del payment_df_filter
        gc.collect()
        
        # 14. Read FSN and add Product Name, FSN columns
        if fsn_file.name.endswith('.csv'):
            FSN = pd.read_csv(fsn_file)
        else:
            FSN = pd.read_excel(fsn_file)
        
        lookup_col = FSN.columns[1]  # Column B → SKU
        return_col = FSN.columns[5]  # Complete Name
        fsn_id_col = FSN.columns[6]  # FSN ID
        
        pivot_df["Seller SKU"] = pivot_df["Seller SKU"].astype(str).str.strip()
        FSN[lookup_col] = FSN[lookup_col].astype(str).str.strip()
        
        fsn_name_lookup = FSN.drop_duplicates(lookup_col).set_index(lookup_col)[return_col]
        fsn_id_lookup = FSN.drop_duplicates(lookup_col).set_index(lookup_col)[fsn_id_col]
        
        pivot_df["Product Name"] = pivot_df["Seller SKU"].map(fsn_name_lookup)
        pivot_df["FSN"] = pivot_df["Seller SKU"].map(fsn_id_lookup)
        
        del FSN
        gc.collect()
        
        # 15. Read PM and add Brand Manager, Brand, Cost, Support
        if pm_file.name.endswith('.csv'):
            PM = pd.read_csv(pm_file)
        else:
            PM = pd.read_excel(pm_file)
        
        pm_lookup_col = PM.columns[0]  # Column A → FSN
        brand_manager_col = PM.columns[4]  # 5th column
        brand_col = PM.columns[5]  # 6th column
        cost_col = PM.columns[8]  # 9th column - Cost
        support_col = PM.columns[9]  # 10th column - Support
        
        pivot_df["FSN"] = pivot_df["FSN"].astype(str).str.strip()
        PM[pm_lookup_col] = PM[pm_lookup_col].astype(str).str.strip()
        
        pm_unique = PM.drop_duplicates(pm_lookup_col).set_index(pm_lookup_col)
        
        pivot_df["Brand Manager"] = pivot_df["FSN"].map(pm_unique[brand_manager_col])
        pivot_df["Brand"] = pivot_df["FSN"].map(pm_unique[brand_col])
        pivot_df["Our Cost Purchase"] = pivot_df["FSN"].map(pm_unique[cost_col])
        pivot_df["Support Amount"] = pivot_df["FSN"].map(pm_unique[support_col])
        
        del PM, pm_unique
        gc.collect()
        
        # 16. Calculate Flipkart Total Fees
        pivot_df["Flipkart Total Fees"] = (
            pivot_df["Sales Amount"]
            - pivot_df["Bank Settlement Value (Rs.) = SUM(J:R)"]
        )
        
        # 17. Calculate Flipkart Fees %
        pivot_df["Flipkart Fees In %"] = np.where(
            pivot_df["Sales Amount"] != 0,
            (pivot_df["Flipkart Total Fees"] / pivot_df["Sales Amount"]) * 100,
            np.nan
        )
        pivot_df["Flipkart Fees In %"] = pivot_df["Flipkart Fees In %"].round(2)
        
        # 18. Fill NA for cost columns
        pivot_df["Our Cost Purchase"] = pivot_df["Our Cost Purchase"].fillna(0)
        pivot_df["Support Amount"] = pivot_df["Support Amount"].fillna(0)
        
        # 19. Calculate Our Cost As Per Qty
        pivot_df["Our Cost As Per Qty"] = (
            pivot_df["Our Cost Purchase"]
            * pivot_df["Sum of Quantity"]
        )
        
        # 20. Calculate Profit
        pivot_df["Profit"] = (
            pivot_df["Bank Settlement Value (Rs.) = SUM(J:R)"]
            - pivot_df["Our Cost As Per Qty"]
        )
        
        # 21. Calculate Profit Percentage
        pivot_df["Profit In Percentage"] = np.where(
            pivot_df["Our Cost As Per Qty"] != 0,
            (pivot_df["Profit"] * 100) / pivot_df["Our Cost As Per Qty"],
            np.nan
        )
        
        # 22. Calculate With BackEnd Price
        pivot_df["With BackEnd Price"] = (
            pivot_df["Our Cost Purchase"]
            - pivot_df["Support Amount"]
        ).round(2)
        
        # 23. Calculate With Support Purchase As Per Qty
        pivot_df["With Support Purchase As Per Qty"] = (
            pivot_df["With BackEnd Price"]
            * pivot_df["Sum of Quantity"]
        )
        
        # 24. Calculate Profit With Support
        pivot_df["Profit With Support"] = (
            pivot_df["Bank Settlement Value (Rs.) = SUM(J:R)"]
            - pivot_df["With Support Purchase As Per Qty"]
        )
        
        # 25. Calculate Profit % With Support
        pivot_df["Profit In Percentage With Support"] = np.where(
            pivot_df["With Support Purchase As Per Qty"] != 0,
            (pivot_df["Profit With Support"] * 100)
            / pivot_df["With Support Purchase As Per Qty"],
            np.nan
        )
        
        # 26. Calculate 3% On Turn Over
        pivot_df["3% On Turn Over"] = (
            pivot_df["Bank Settlement Value (Rs.) = SUM(J:R)"] * 0.03
        )
        
        # 27. Calculate After 3% Profit
        pivot_df["After 3% Profit"] = (
            pivot_df["Profit With Support"]
            - pivot_df["3% On Turn Over"]
        )
        
        # 28. Calculate After 3% Percentage
        pivot_df["After 3%"] = np.where(
            pivot_df["With Support Purchase As Per Qty"] != 0,
            (pivot_df["After 3% Profit"] * 100)
            / pivot_df["With Support Purchase As Per Qty"],
            np.nan
        )
        
        # 29. Filter by profit margin > -45%
        pivot_df["Profit In Percentage With Support"] = pd.to_numeric(
            pivot_df["Profit In Percentage With Support"],
            errors="coerce"
        )
        pivot_df = pivot_df[pivot_df["Profit In Percentage With Support"] > -45]
        
        # 30. Create Brand Pivot
        brand_pivot = pd.pivot_table(
            pivot_df,
            index="Brand",
            values="After 3% Profit",
            aggfunc="sum",
            margins=True,
            margins_name="Grand Total"
        ).reset_index()
        
        # 31. Add Grand Total row
        numeric_columns = [
            "Bank Settlement Value (Rs.) = SUM(J:R)", "Sales Amount",
            "Sum of Quantity", "Flipkart Total Fees", "Flipkart Fees In %",
            "Our Cost Purchase", "Our Cost As Per Qty", "Profit",
            "Profit In Percentage", "Support Amount", "With BackEnd Price",
            "With Support Purchase As Per Qty", "Profit With Support",
            "Profit In Percentage With Support", "3% On Turn Over",
            "After 3% Profit", "After 3%"
        ]
        
        grand_total_row = {col: "" for col in pivot_df.columns}
        for col in numeric_columns:
            if col in pivot_df.columns:
                pivot_df[col] = pd.to_numeric(pivot_df[col], errors="coerce")
                grand_total_row[col] = pivot_df[col].sum()
        
        grand_total_row["Order Date"] = "Grand Total"
        pivot_df = pd.concat(
            [pivot_df, pd.DataFrame([grand_total_row])],
            ignore_index=True
        )
        
        # Reorder columns to match requested sequence
        column_order = [
            "Order Date",
            "Order item ID",
            "Order ID",
            "Seller SKU",
            "Brand Manager",
            "Brand",
            "Product Name",
            "FSN",
            "Sum of Quantity",
            "Sales Amount",
            "Flipkart Total Fees",
            "Flipkart Fees In %",
            "Bank Settlement Value (Rs.) = SUM(J:R)",
            "Our Cost Purchase",
            "Our Cost As Per Qty",
            "Profit",
            "Profit In Percentage",
            "Support Amount",
            "With BackEnd Price",
            "With Support Purchase As Per Qty",
            "Profit With Support",
            "Profit In Percentage With Support",
            "3% On Turn Over",
            "After 3% Profit",
            "After 3%"
        ]
        
        # Only include columns that exist in the dataframe
        final_columns = [col for col in column_order if col in pivot_df.columns]
        pivot_df = pivot_df[final_columns]
        
        return pivot_df, brand_pivot
    
    except Exception as e:
        raise Exception(f"Processing error: {str(e)}")


# Main processing
if all([payment_file, uni_orders_file, fsn_file, pm_file]):
    # Check if files changed
    current_hash = get_files_hash([payment_file, uni_orders_file, fsn_file, pm_file])
    
    # Process only if files are new or changed
    if st.session_state.files_hash != current_hash or st.session_state.processed_data is None:
        with st.spinner("Processing files... Please wait."):
            try:
                # Reset file positions before reading
                payment_file.seek(0)
                uni_orders_file.seek(0)
                fsn_file.seek(0)
                pm_file.seek(0)
                
                pivot_df, brand_pivot = process_data(
                    payment_file, uni_orders_file, fsn_file, pm_file
                )
                
                # Store in session state
                st.session_state.processed_data = pivot_df
                st.session_state.brand_pivot = brand_pivot
                st.session_state.files_hash = current_hash
                
                # Save to MongoDB
                try:
                    save_reconciliation_report(
                        collection_name="flipkart_pl",
                        invoice_no=f"FLIPKART_PL_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                        summary_data={
                            "total_sales": float(pivot_df["Sales Amount"].sum()) if "Sales Amount" in pivot_df.columns else 0,
                            "total_profit": float(pivot_df["After 3% Profit"].sum()) if "After 3% Profit" in pivot_df.columns else 0,
                            "total_rows": len(pivot_df)
                        },
                        line_items_data=pivot_df,
                        metadata={"report_type": "flipkart_pl"}
                    )
                except Exception:
                    pass
                
                gc.collect()
                
            except Exception as e:
                st.error(f"❌ Error processing files: {str(e)}")
                st.info("Please ensure all files are in the correct format as expected.")
                st.stop()
    
    # Use data from session state
    pivot_df = st.session_state.processed_data
    brand_pivot = st.session_state.brand_pivot
    
    if pivot_df is not None:
        try:
            # Calculate summary metrics (excluding Grand Total row)
            data_df = pivot_df[pivot_df["Order Date"] != "Grand Total"].copy()
            
            total_sales = pd.to_numeric(data_df["Sales Amount"], errors="coerce").sum()
            total_settlement = pd.to_numeric(data_df["Bank Settlement Value (Rs.) = SUM(J:R)"], errors="coerce").sum()
            total_quantity = pd.to_numeric(data_df["Sum of Quantity"], errors="coerce").sum()
            total_profit = pd.to_numeric(data_df["After 3% Profit"], errors="coerce").sum()
            total_flipkart_fees = pd.to_numeric(data_df["Flipkart Total Fees"], errors="coerce").sum()
            
            # Display metrics
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("📈 Total Sales", f"₹{total_sales:,.2f}")
            with col2:
                st.metric("🏦 Settlement Value", f"₹{total_settlement:,.2f}")
            with col3:
                st.metric("📦 Total Quantity", f"{int(total_quantity):,}")
            with col4:
                st.metric("💰 Flipkart Fees", f"₹{total_flipkart_fees:,.2f}")
            with col5:
                profit_color = "normal" if total_profit >= 0 else "inverse"
                st.metric("🎯 Net Profit (After 3%)", f"₹{total_profit:,.2f}", delta_color=profit_color)
            
            st.markdown("---")
            
            # Tabs for different views
            tab1, tab2 = st.tabs(["📋 Detailed P&L Report", "📊 Brand-wise Summary"])
            
            with tab1:
                st.subheader("Detailed P&L Report")
                styled_df = style_dataframe(pivot_df)
                st.dataframe(
                    styled_df,
                    use_container_width=True,
                    height=500
                )
                
                # Use base64 download links instead of st.download_button
                st.markdown("### 📥 Download Options")
                col1, col2 = st.columns(2)
                with col1:
                    excel_link = get_download_link(pivot_df, "Flipkart_PL_Report.xlsx", "excel")
                    st.markdown(excel_link, unsafe_allow_html=True)
                with col2:
                    csv_link = get_download_link(pivot_df, "Flipkart_PL_Report.csv", "csv")
                    st.markdown(csv_link, unsafe_allow_html=True)
            
            with tab2:
                st.subheader("Brand-wise Profit Summary")
                st.dataframe(
                    brand_pivot,
                    use_container_width=True,
                    height=400
                )
                
                st.markdown("### 📥 Download Options")
                col1, col2 = st.columns(2)
                with col1:
                    excel_link = get_download_link(brand_pivot, "Flipkart_Brand_Summary.xlsx", "excel")
                    st.markdown(excel_link, unsafe_allow_html=True)
                with col2:
                    csv_link = get_download_link(brand_pivot, "Flipkart_Brand_Summary.csv", "csv")
                    st.markdown(csv_link, unsafe_allow_html=True)
            
            st.success(f"✅ Successfully processed {len(data_df)} orders!")
            
        except Exception as e:
            st.error(f"❌ Error displaying data: {str(e)}")

else:
    st.info("👈 Please upload all 4 required files in the sidebar to begin analysis.")
    
    st.markdown("""
    ### Required Files:
    1. **Payment File Flipkart.xlsx** - Payment file with 'Orders' sheet
    2. **Uni Orders.xlsx** - Unique orders reference file
    3. **FSN.xlsx** - Product FSN master file
    4. **flipkart PM.xlsx** - Purchase master with cost and support data
    """)