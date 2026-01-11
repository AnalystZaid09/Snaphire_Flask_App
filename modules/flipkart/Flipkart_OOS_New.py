import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from common.ui_utils import (
    apply_professional_style, 
    get_download_filename, 
    render_header,
    download_module_report
)

MODULE_NAME = "flipkart"

# Page configuration
st.set_page_config(
    page_title="Flipkart OOS Inventory Management",
    page_icon="📦",
    layout="wide"
)
apply_professional_style()

# Helper Functions
def normalize_sku(series):
    return series.astype(str)

def get_doc_color_hex(doc_value):
    """Return hex color based on DOC value"""
    if pd.isna(doc_value) or doc_value == 0 or doc_value == '':
        return None
    
    try:
        doc_val = float(doc_value)
    except:
        return None
    
    if doc_val < 7:
        return 'FF4444'  # Red
    elif doc_val < 15:
        return 'FF8800'  # Orange
    elif doc_val < 30:
        return '44FF44'  # Green
    elif doc_val < 45:
        return 'FFFF44'  # Yellow
    elif doc_val < 60:
        return '44DDFF'  # Sky Blue
    elif doc_val < 90:
        return '8B4513'  # Brown
    else:
        return '000000'  # Black

def get_doc_font_color(doc_value):
    """Return font color based on DOC value"""
    if pd.isna(doc_value) or doc_value == 0 or doc_value == '':
        return '000000'
    
    try:
        doc_val = float(doc_value)
    except:
        return '000000'
    
    if doc_val < 7 or doc_val < 15 or doc_val >= 90 or (doc_val >= 60 and doc_val < 90):
        return 'FFFFFF'
    else:
        return '000000'

def get_doc_color(doc_value):
    """Return color based on DOC value for Streamlit display"""
    if pd.isna(doc_value) or doc_value == 0 or doc_value == '':
        return ''
    
    try:
        doc_val = float(doc_value)
    except:
        return ''
    
    if doc_val < 7:
        return 'background-color: #ff4444; color: white;'
    elif doc_val < 15:
        return 'background-color: #ff8800; color: white;'
    elif doc_val < 30:
        return 'background-color: #44ff44; color: black;'
    elif doc_val < 45:
        return 'background-color: #ffff44; color: black;'
    elif doc_val < 60:
        return 'background-color: #44ddff; color: black;'
    elif doc_val < 90:
        return 'background-color: #8b4513; color: white;'
    else:
        return 'background-color: #000000; color: white;'

def style_doc_column(df):
    """Apply color styling to DOC column"""
    def apply_color(val):
        return get_doc_color(val)
    
    styled = df.style.applymap(apply_color, subset=['DOC'])
    return styled

def to_excel(df, apply_doc_formatting=False):
    """Convert dataframe to Excel with optional DOC formatting"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        if apply_doc_formatting and 'DOC' in df.columns:
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            doc_col_idx = df.columns.get_loc('DOC') + 1
            
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=doc_col_idx)
                doc_value = df.iloc[row_idx - 2]['DOC']
                
                bg_color = get_doc_color_hex(doc_value)
                font_color = get_doc_font_color(doc_value)
                
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    cell.font = Font(color=font_color, bold=True)
    
    return output.getvalue()

def create_stock_pivot(df, id_column="Product Id"):
    """Creates pivot table with Brand, Product ID, and sums of DOC, DRR, CP"""
    df_copy = df.copy()
    for col in ["DOC", "DRR", "CP"]:
        if col in df_copy.columns:
            df_copy[col] = pd.to_numeric(df_copy[col], errors="coerce").fillna(0)
    
    # Determine which ID column to use
    if id_column not in df_copy.columns:
        if "Flipkart's Identifier of the product" in df_copy.columns:
            id_column = "Flipkart's Identifier of the product"
        elif "Product Id" in df_copy.columns:
            id_column = "Product Id"
    
    pivot = pd.pivot_table(
        df_copy,
        index=["Brand", id_column],
        values=["DOC", "DRR", "CP"],
        aggfunc="sum",
        margins=True,
        margins_name="Grand Total"
    )
    pivot = pivot.reset_index()
    pivot.rename(
        columns={
            "DOC": "Sum of DOC",
            "DRR": "Sum of DRR",
            "CP": "Sum of CP"
        },
        inplace=True
    )
    return pivot

def process_business_report(business_file, purchase_master_file, inventory_file, no_of_days, doc_threshold):
    """Process Flipkart business report"""
    
    # Read Business Report
    if business_file.name.endswith('.csv'):
        Business_Report = pd.read_csv(business_file)
    else:
        Business_Report = pd.read_excel(business_file)
    
    # Filter only items with Final Sale Units > 0
    Business_Report = Business_Report[Business_Report["Final Sale Units"] > 0].reset_index(drop=True)
    
    # Create Business Pivot
    Business_Pivot = pd.pivot_table(
        Business_Report,
        index=["Product Id", "SKU ID"],
        values=["Gross Units", "GMV", "Final Sale Units", "Final Sale Amount"],
        aggfunc="sum",
        margins=True,
        margins_name="Grand Total"
    )
    Business_Pivot = Business_Pivot.reset_index()
    Business_Pivot.rename(columns={"Final Sale Units": "Sum of Final Sale Units"}, inplace=True)
    
    # Sort by Final Sale Units
    Business_Pivot_no_total = Business_Pivot[Business_Pivot["Product Id"] != "Grand Total"]
    Business_Pivot_total = Business_Pivot[Business_Pivot["Product Id"] == "Grand Total"]
    Business_Pivot = pd.concat([
        Business_Pivot_no_total.sort_values("Sum of Final Sale Units", ascending=False),
        Business_Pivot_total
    ], ignore_index=True)
    
    # Read Purchase Master
    if purchase_master_file.name.endswith('.csv'):
        purchase_master = pd.read_csv(purchase_master_file)
    else:
        purchase_master = pd.read_excel(purchase_master_file)
    
    purchase_master["FNS"] = normalize_sku(purchase_master["FNS"])
    
    # Map Purchase Master data
    purchase_master1 = purchase_master[["FNS", "Vendor SKU Codes", "Brand Manager", "Brand", "Product Name"]].copy()
    purchase_master1 = purchase_master1.drop_duplicates(subset="FNS", keep="first")
    
    vendor_sku_map = purchase_master1.set_index("FNS")["Vendor SKU Codes"]
    brand_map = purchase_master1.set_index("FNS")["Brand"]
    product_map = purchase_master1.set_index("FNS")["Product Name"]
    manager_map = purchase_master1.set_index("FNS")["Brand Manager"]
    
    Business_Pivot["Vendor SKU Codes"] = Business_Pivot["Product Id"].map(vendor_sku_map)
    Business_Pivot["Brand"] = Business_Pivot["Product Id"].map(brand_map)
    Business_Pivot["Product Name"] = Business_Pivot["Product Id"].map(product_map)
    Business_Pivot["Brand Manager"] = Business_Pivot["Product Id"].map(manager_map)
    
    # Map CP
    purchase_master2 = purchase_master.copy()
    purchase_master2 = purchase_master2.drop_duplicates(subset="FNS", keep="first").set_index("FNS")
    purchase_master2["CP"] = pd.to_numeric(purchase_master2["CP"], errors="coerce")
    Business_Pivot["CP"] = Business_Pivot["Product Id"].map(purchase_master2["CP"]).fillna(0)
    
    # Reorder columns
    Business_Pivot = Business_Pivot[[
        "Product Id", "SKU ID", "Vendor SKU Codes", "Brand", "Product Name",
        "Brand Manager", "GMV", "Gross Units", "Final Sale Amount", "Sum of Final Sale Units", "CP"
    ]]
    
    Business_Pivot.fillna("", inplace=True)
    
    # Calculate As Per Qty CP
    Business_Pivot["CP_numeric"] = pd.to_numeric(Business_Pivot["CP"], errors="coerce").fillna(0)
    Business_Pivot["As Per Qty CP"] = Business_Pivot["Sum of Final Sale Units"] * Business_Pivot["CP_numeric"]
    Business_Pivot.drop("CP_numeric", axis=1, inplace=True)
    
    # Calculate DRR
    Business_Pivot["DRR"] = (Business_Pivot["Sum of Final Sale Units"] / no_of_days).round(2)
    
    # Read Inventory
    if inventory_file.name.endswith('.csv'):
        Inventory = pd.read_csv(inventory_file, header=1)
    else:
        Inventory = pd.read_excel(inventory_file, header=1)
    
    # Create Inventory Pivot
    inventory_pivot = pd.pivot_table(
        Inventory,
        index="Flipkart's Identifier of the product",
        values="Current stock count for your product",
        aggfunc="sum",
        margins=True,
        margins_name="Grand Total"
    )
    inventory_pivot = inventory_pivot.reset_index()
    inventory_pivot.rename(
        columns={"Flipkart's Identifier of the product": "FNS", 
                 "Current stock count for your product": "Current Stock"},
        inplace=True
    )
    
    # Map inventory to Business Pivot
    inventory_lookup = inventory_pivot.drop_duplicates(subset="FNS", keep="first").set_index("FNS")
    Business_Pivot["Current Stock"] = Business_Pivot["Product Id"].map(inventory_lookup["Current Stock"]).fillna(0)
    
    # Calculate DOC
    Business_Pivot["DOC"] = Business_Pivot["Current Stock"] / Business_Pivot["DRR"]
    Business_Pivot["DOC"] = Business_Pivot["DOC"].replace([np.inf, -np.inf], np.nan)
    Business_Pivot.loc[Business_Pivot["Product Id"] == "Grand Total", "DOC"] = ""
    Business_Pivot["DOC"] = Business_Pivot["DOC"].apply(
        lambda x: round(x, 2) if x != "" and pd.notna(x) else ""
    )
    
    # Create OOS Report
    OOS_Report = Business_Pivot[Business_Pivot["Product Id"] != "Grand Total"].copy()
    OOS_Report = OOS_Report[OOS_Report["Current Stock"] == 0].reset_index(drop=True)
    OOS_Report["CP"] = pd.to_numeric(OOS_Report["CP"], errors="coerce").fillna(0)
    OOS_Report["DOC"] = pd.to_numeric(OOS_Report["DOC"], errors="coerce").fillna(0)
    OOS_Report["DRR"] = pd.to_numeric(OOS_Report["DRR"], errors="coerce").fillna(0)
    
    # Create Overstock Report
    Overstock_Report = Business_Pivot[Business_Pivot["Product Id"] != "Grand Total"].copy()
    Overstock_Report["DOC"] = pd.to_numeric(Overstock_Report["DOC"], errors="coerce").fillna(0)
    Overstock_Report = Overstock_Report[Overstock_Report["DOC"] > doc_threshold].reset_index(drop=True)
    
    # Create pivots
    OOS_Pivot = create_stock_pivot(OOS_Report, id_column="Product Id")
    Overstock_Pivot = create_stock_pivot(Overstock_Report, id_column="Product Id")
    
    return Business_Pivot, OOS_Report, Overstock_Report, OOS_Pivot, Overstock_Pivot

def process_inventory_report(inventory_file, purchase_master_file, business_pivot, no_of_days_inventory, doc_inventory_threshold):
    """Process inventory report"""
    
    # Read Inventory
    if inventory_file.name.endswith('.csv'):
        Inventory = pd.read_csv(inventory_file, header=1)
    else:
        Inventory = pd.read_excel(inventory_file, header=1)
    
    # Create Inventory Report Pivot
    Inventory_Report_Pivot = pd.pivot_table(
        Inventory,
        index="Flipkart's Identifier of the product",
        values="Current stock count for your product",
        aggfunc="sum"
    )
    Inventory_Report_Pivot = Inventory_Report_Pivot.reset_index()
    
    # Read Purchase Master
    if purchase_master_file.name.endswith('.csv'):
        purchase_master = pd.read_csv(purchase_master_file)
    else:
        purchase_master = pd.read_excel(purchase_master_file)
    
    purchase_master["FNS"] = normalize_sku(purchase_master["FNS"])
    
    # Map Purchase Master data
    pm_lookup = purchase_master[["FNS", "Vendor SKU Codes", "Brand", "Brand Manager", "Product Name", "CP"]].copy()
    pm_lookup = pm_lookup.drop_duplicates(subset="FNS", keep="first").set_index("FNS")
    
    Inventory_Report_Pivot["Vendor SKU Codes"] = Inventory_Report_Pivot["Flipkart's Identifier of the product"].map(pm_lookup["Vendor SKU Codes"])
    Inventory_Report_Pivot["Brand"] = Inventory_Report_Pivot["Flipkart's Identifier of the product"].map(pm_lookup["Brand"])
    Inventory_Report_Pivot["Brand Manager"] = Inventory_Report_Pivot["Flipkart's Identifier of the product"].map(pm_lookup["Brand Manager"])
    Inventory_Report_Pivot["Product Name"] = Inventory_Report_Pivot["Flipkart's Identifier of the product"].map(pm_lookup["Product Name"])
    Inventory_Report_Pivot["CP"] = Inventory_Report_Pivot["Flipkart's Identifier of the product"].map(pm_lookup["CP"])
    
    # Reorder columns
    Inventory_Report_Pivot = Inventory_Report_Pivot[[
        "Flipkart's Identifier of the product", "Vendor SKU Codes", "Brand", "Brand Manager",
        "Product Name", "Current stock count for your product", "CP"
    ]]
    
    # Map Final Sales Units from Business Pivot
    business_lookup = business_pivot[["Product Id", "Sum of Final Sale Units"]].copy()
    business_lookup = business_lookup.drop_duplicates(subset="Product Id", keep="first").set_index("Product Id")
    Inventory_Report_Pivot["Final Sales Units"] = Inventory_Report_Pivot["Flipkart's Identifier of the product"].map(
        business_lookup["Sum of Final Sale Units"]
    )
    
    # Fill NaN values
    Inventory_Report_Pivot[["Final Sales Units","CP"]] = Inventory_Report_Pivot[["Final Sales Units","CP"]].fillna(0)
    
    # Clean CP
    Inventory_Report_Pivot["CP"] = (
        Inventory_Report_Pivot["CP"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    Inventory_Report_Pivot["CP"] = pd.to_numeric(Inventory_Report_Pivot["CP"], errors="coerce").round(2)
    
    # Calculate As Per Qty CP
    Inventory_Report_Pivot["As Per Qty CP"] = Inventory_Report_Pivot["CP"] * Inventory_Report_Pivot["Final Sales Units"]
    
    # Calculate DRR
    Inventory_Report_Pivot["DRR"] = (Inventory_Report_Pivot["Final Sales Units"] / no_of_days_inventory).round(2)
    
    # Calculate DOC
    Inventory_Report_Pivot["DOC"] = (
        Inventory_Report_Pivot["Current stock count for your product"] / Inventory_Report_Pivot["DRR"]
    )
    Inventory_Report_Pivot["DOC"] = Inventory_Report_Pivot["DOC"].replace([np.inf, -np.inf], 0).round(2)
    Inventory_Report_Pivot["DOC"] = pd.to_numeric(Inventory_Report_Pivot["DOC"], errors="coerce").fillna(0)
    
    # Filter out items with no sales
    #Inventory_Report_Pivot = Inventory_Report_Pivot[Inventory_Report_Pivot["Final Sales Units"] != 0]
    # Instead of removing, set Final Sales Units negative value to 0
    Inventory_Report_Pivot["Final Sales Units"] = Inventory_Report_Pivot["Final Sales Units"].clip(lower=0)

    Inventory_Report_Pivot = Inventory_Report_Pivot[[
        "Flipkart's Identifier of the product", "Vendor SKU Codes", "Brand", "Brand Manager",
        "Product Name", "Current stock count for your product",
        "Final Sales Units",
        "CP", "As Per Qty CP", "DRR", "DOC" # keep your metric columns
    ]]
    
    # Create OOS Inventory
    OOS_Inventory = Inventory_Report_Pivot[
        Inventory_Report_Pivot["Current stock count for your product"] == 0
    ].reset_index(drop=True)
    
    # Create Overstock Inventory
    Overstock_Inventory = Inventory_Report_Pivot[
        Inventory_Report_Pivot["DOC"] >= doc_inventory_threshold
    ].reset_index(drop=True)
    
    # Create pivots
    OOS_Inventory_Pivot = create_stock_pivot(OOS_Inventory, id_column="Flipkart's Identifier of the product")
    Overstock_Inventory_Pivot = create_stock_pivot(Overstock_Inventory, id_column="Flipkart's Identifier of the product")
    
    return Inventory_Report_Pivot, OOS_Inventory, Overstock_Inventory, OOS_Inventory_Pivot, Overstock_Inventory_Pivot

# Main App
render_header("Flipkart OOS Inventory Management System")

# Sidebar for file uploads
st.sidebar.header("📁 Upload Files")

business_file = st.sidebar.file_uploader("Business Report (Earn More Report)", type=['csv', 'xlsx'])
purchase_master_file = st.sidebar.file_uploader("Purchase Master (Flipkart PM)", type=['csv', 'xlsx'])
inventory_file = st.sidebar.file_uploader("Manage Inventory (Listing Report)", type=['csv', 'xlsx', 'xls'])

st.sidebar.header("⚙️ Parameters")
no_of_days = st.sidebar.number_input("Number of Days (Business)", min_value=1, value=30)
doc_threshold = st.sidebar.number_input("DOC Threshold (Business)", min_value=0, value=30)
no_of_days_inventory = st.sidebar.number_input("Number of Days (Inventory)", min_value=1, value=30)
doc_inventory_threshold = st.sidebar.number_input("DOC Threshold (Inventory)", min_value=0, value=30)

# DOC Legend
st.sidebar.markdown("---")
st.sidebar.markdown("### 🎨 DOC Color Legend")
st.sidebar.markdown("🔴 **Red (0-7)**: Critical")
st.sidebar.markdown("🟠 **Orange (7-15)**: Low")
st.sidebar.markdown("🟢 **Green (15-30)**: Optimal")
st.sidebar.markdown("🟡 **Yellow (30-45)**: Monitor")
st.sidebar.markdown("🔵 **Sky Blue (45-60)**: High")
st.sidebar.markdown("🟤 **Brown (60-90)**: Excess")
st.sidebar.markdown("⬛ **Black (>90)**: Overstocked")

# Process data when all required files are uploaded
if business_file and purchase_master_file and inventory_file:
    try:
        import io
        
        # Reset file pointers
        business_file.seek(0)
        purchase_master_file.seek(0)
        inventory_file.seek(0)
        
        # Read files into memory
        business_bytes = business_file.read()
        purchase_bytes = purchase_master_file.read()
        inventory_bytes = inventory_file.read()
        
        # Create new file-like objects
        business_io = io.BytesIO(business_bytes)
        purchase_io = io.BytesIO(purchase_bytes)
        inventory_io = io.BytesIO(inventory_bytes)
        
        # Set names
        business_io.name = business_file.name
        purchase_io.name = purchase_master_file.name
        inventory_io.name = inventory_file.name
        
        # Process Business Report
        Business_Pivot, OOS_Report, Overstock_Report, OOS_Pivot, Overstock_Pivot = process_business_report(
            business_io, purchase_io, inventory_io,
            no_of_days, doc_threshold
        )
        
        # Reset file pointers
        purchase_io.seek(0)
        inventory_io.seek(0)
        
        # Process Inventory Report
        Inventory_Report_Pivot, OOS_Inventory, Overstock_Inventory, OOS_Inventory_Pivot, Overstock_Inventory_Pivot = process_inventory_report(
            inventory_io, purchase_io, Business_Pivot,
            no_of_days_inventory, doc_inventory_threshold
        )
        
        # Create tabs
        tab1, tab2 = st.tabs(["📊 Business Report", "📦 Inventory Report"])
        
        # Tab 1: Business Report
        with tab1:
            st.header("Business Report")
            
            sub_tab1, sub_tab2, sub_tab3 = st.tabs(["Main Report", "OOS Report", "Overstock Report"])
            
            with sub_tab1:
                st.subheader("Business Pivot Report")
                
                if 'DOC' in Business_Pivot.columns:
                    styled_df = style_doc_column(Business_Pivot)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Business_Pivot, use_container_width=True, height=600)
                
                download_module_report(
                    df=Business_Pivot,
                    module_name=MODULE_NAME,
                    report_name="Business Pivot",
                    button_label="📥 Download Business Pivot",
                    key="dl_fk_oos_biz_pivot"
                )
            
            with sub_tab2:
                st.subheader("Out of Stock (OOS) Report")
                
                if 'DOC' in OOS_Report.columns:
                    styled_df = style_doc_column(OOS_Report)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(OOS_Report, use_container_width=True, height=600)
                
                download_module_report(
                    df=OOS_Report,
                    module_name=MODULE_NAME,
                    report_name="OOS Report",
                    button_label="📥 Download OOS Report",
                    key="dl_fk_oos_report"
                )
                
                st.subheader("OOS Pivot Table")
                st.dataframe(OOS_Pivot, use_container_width=True)
                download_module_report(
                    df=OOS_Pivot,
                    module_name=MODULE_NAME,
                    report_name="OOS Pivot",
                    button_label="📥 Download OOS Pivot",
                    key="dl_fk_oos_pivot"
                )
            
            with sub_tab3:
                st.subheader("Overstock Report")
                
                if 'DOC' in Overstock_Report.columns:
                    styled_df = style_doc_column(Overstock_Report)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Overstock_Report, use_container_width=True, height=600)
                
                download_module_report(
                    df=Overstock_Report,
                    module_name=MODULE_NAME,
                    report_name="Overstock Report",
                    button_label="📥 Download Overstock Report",
                    key="dl_fk_overstock_report"
                )
                
                st.subheader("Overstock Pivot Table")
                st.dataframe(Overstock_Pivot, use_container_width=True)
                download_module_report(
                    df=Overstock_Pivot,
                    module_name=MODULE_NAME,
                    report_name="Overstock Pivot",
                    button_label="📥 Download Overstock Pivot",
                    key="dl_fk_overstock_pivot"
                )
        
        # Tab 2: Inventory Report
        with tab2:
            st.header("Inventory Report")
            
            sub_tab1, sub_tab2, sub_tab3 = st.tabs(["Main Report", "OOS Inventory", "Overstock Inventory"])
            
            with sub_tab1:
                st.subheader("Inventory Report Pivot")
                
                if 'DOC' in Inventory_Report_Pivot.columns:
                    styled_df = style_doc_column(Inventory_Report_Pivot)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Inventory_Report_Pivot, use_container_width=True, height=600)
                
                download_module_report(
                    df=Inventory_Report_Pivot,
                    module_name=MODULE_NAME,
                    report_name="Inventory Report",
                    button_label="📥 Download Inventory Report",
                    key="dl_fk_inv_report"
                )
            
            with sub_tab2:
                st.subheader("OOS Inventory Report")
                
                if 'DOC' in OOS_Inventory.columns:
                    styled_df = style_doc_column(OOS_Inventory)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(OOS_Inventory, use_container_width=True, height=600)
                
                download_module_report(
                    df=OOS_Inventory,
                    module_name=MODULE_NAME,
                    report_name="OOS Inventory",
                    button_label="📥 Download OOS Inventory",
                    key="dl_fk_oos_inv"
                )
                
                st.subheader("OOS Inventory Pivot Table")
                st.dataframe(OOS_Inventory_Pivot, use_container_width=True)
                download_module_report(
                    df=OOS_Inventory_Pivot,
                    module_name=MODULE_NAME,
                    report_name="OOS Inventory Pivot",
                    button_label="📥 Download OOS Inventory Pivot",
                    key="dl_fk_oos_inv_pivot"
                )
            
            with sub_tab3:
                st.subheader("Overstock Inventory Report")
                
                if 'DOC' in Overstock_Inventory.columns:
                    styled_df = style_doc_column(Overstock_Inventory)
                    st.dataframe(styled_df, use_container_width=True, height=600)
                else:
                    st.dataframe(Overstock_Inventory, use_container_width=True, height=600)
                
                download_module_report(
                    df=Overstock_Inventory,
                    module_name=MODULE_NAME,
                    report_name="Overstock Inventory",
                    button_label="📥 Download Overstock Inventory",
                    key="dl_fk_over_inv"
                )
                
                st.subheader("Overstock Inventory Pivot Table")
                st.dataframe(Overstock_Inventory_Pivot, use_container_width=True)
                download_module_report(
                    df=Overstock_Inventory_Pivot,
                    module_name=MODULE_NAME,
                    report_name="Overstock Inventory Pivot",
                    button_label="📥 Download Overstock Inventory Pivot",
                    key="dl_fk_over_inv_pivot"
                )
            
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        st.exception(e)
else:
    st.info("👆 Please upload all required files (Business Report, Purchase Master, and Manage Inventory) to begin.")
    
    # Display sample information
    with st.expander("ℹ️ File Requirements"):
        st.markdown("""
        ### Required Files:
        
        1. **Business Report (Earn More Report)**
           - Should contain columns: Product Id, SKU ID, Final Sale Units, GMV, etc.
           - Filters: Only items with Final Sale Units > 0
        
        2. **Purchase Master (Flipkart PM)**
           - Should contain columns: FNS, Vendor SKU Codes, Brand, Brand Manager, Product Name, CP
           - Used for mapping product details and cost price
        
        3. **Manage Inventory (Listing Report)**
           - Should contain: Flipkart's Identifier of the product, Current stock count
           - Header starts at row 2 (skip first row)
        
        ### Key Metrics:
        - **DRR (Daily Run Rate)**: Final Sale Units / Number of Days
        - **DOC (Days of Coverage)**: Current Stock / DRR
        - **OOS (Out of Stock)**: Items with Current Stock = 0
        - **Overstock**: Items with DOC > Threshold
        """)
