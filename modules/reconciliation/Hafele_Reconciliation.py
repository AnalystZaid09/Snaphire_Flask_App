# app.py
import streamlit as st
import pandas as pd
import io
import re
import os
from decimal import Decimal
from dotenv import load_dotenv
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from common.mongo import save_reconciliation_report
from common.ui_utils import (
    apply_professional_style, 
    get_download_filename, 
    render_header,
    download_module_report
)

MODULE_NAME = "reconciliation"

# Load .env for local dev (optional)
load_dotenv()

# Server-side Azure credentials (must be set)
AZURE_ENDPOINT = os.getenv("AZURE_ENDPOINT")
AZURE_KEY = os.getenv("AZURE_KEY")

# Apply Professional UI
apply_professional_style()
render_header("Hafele Invoice ⇄ Excel Reconciliation", None)
st.markdown(
    "Upload one invoice **PDF** and one **Excel** (items/PO sheet). "
)

# ---------- Inputs ----------
left, right = st.columns([2, 1])
with left:
    uploaded_pdf = st.file_uploader("Upload invoice PDF", type=["pdf"])
    uploaded_excel = st.file_uploader("Upload Excel (.xlsx/.xls)", type=["xlsx", "xls"])
    header_row_idx = st.number_input("Header row index (0-based). Use 1 for second row as header", value=1, min_value=0, max_value=10)
with right:
    st.markdown("**Options**")
    show_only_mismatches = st.checkbox("Show only mismatches in table", value=False)
    abs_tol = st.number_input("Amount absolute tolerance", value=1.0, step=0.5)
    rel_tol = st.number_input("Amount relative tolerance (fraction)", value=0.005, step=0.001, format="%.4f")
run_button = st.button("Run Reconciliation")

# ---------- Helpers ----------
def normalize_sku_for_match(sku):
    if sku is None:
        return ""
    s = str(sku).strip()
    s = s.replace("HA-", "").replace("ha-", "")
    s = re.sub(r'[^0-9A-Za-z]', '', s)
    return s.lower()

def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return float(v)
        except Exception:
            return None
    s = str(v).strip()
    if s == "" or s.lower() in {"nan","na","-","--","..."}:
        return None
    cleaned = re.sub(r"[^\d\.\-]", "", s.replace(",", ""))
    if cleaned in {"", ".", "-"}:
        return None
    try:
        return float(cleaned)
    except Exception:
        try:
            return float(s.replace(",", ""))
        except Exception:
            return None

def amounts_close(a, b, abs_tol=1.0, rel_tol=0.005):
    if a is None or b is None:
        return False
    try:
        a_f = float(a); b_f = float(b)
    except Exception:
        return False
    diff = abs(a_f - b_f)
    if diff <= abs_tol:
        return True
    denom = max(abs(a_f), abs(b_f), 1e-9)
    return (diff / denom) <= rel_tol

def parse_invoice_from_bytes(client, pdf_bytes):
    poller = client.begin_analyze_document(model_id="prebuilt-invoice", body=pdf_bytes)
    result = poller.result()
    if not result.documents:
        return None
    doc = result.documents[0]
    fields = doc.fields or {}

    def safe_field(field):
        if field is None:
            return None
        if hasattr(field, "value_string") and field.value_string is not None:
            return field.value_string
        if hasattr(field, "value_number") and field.value_number is not None:
            return field.value_number
        if hasattr(field, "value_date") and field.value_date is not None:
            return field.value_date
        if hasattr(field, "value_currency") and field.value_currency is not None:
            cur = field.value_currency
            amt = None
            try:
                amt = getattr(cur, "amount", None)
            except Exception:
                try:
                    amt = cur.get("amount")
                except Exception:
                    amt = None
            return {"amount": to_number(amt), "currency": getattr(cur, "currency", getattr(cur, "currencyCode", None))}
        if hasattr(field, "value_array") and field.value_array is not None:
            return field.value_array
        if hasattr(field, "value_object") and field.value_object is not None:
            return field.value_object
        try:
            return str(field)
        except:
            return None

    pdf_po = safe_field(fields.get("PurchaseOrder"))
    pdf_inv_id = safe_field(fields.get("InvoiceId"))
    inv_total_field = safe_field(fields.get("InvoiceTotal"))
    pdf_invoice_total = None
    if isinstance(inv_total_field, dict) and "amount" in inv_total_field:
        pdf_invoice_total = inv_total_field["amount"]
    else:
        pdf_invoice_total = to_number(inv_total_field)

    items = []
    items_field = fields.get("Items")
    if items_field is not None:
        arr = getattr(items_field, "value_array", None)
        if arr is None:
            arr = safe_field(items_field)
        if isinstance(arr, list):
            for item_field in arr:
                obj = getattr(item_field, "value_object", None)
                if obj is None:
                    obj = safe_field(item_field)
                if isinstance(obj, dict):
                    def get_obj_field(k):
                        f = obj.get(k) if isinstance(obj, dict) else None
                        return safe_field(f) if f is not None else None
                    prod_code = get_obj_field("ProductCode") or get_obj_field("Product Code") or get_obj_field("productcode")
                    desc = get_obj_field("Description") or get_obj_field("description") or get_obj_field("Name")
                    qty = get_obj_field("Quantity")
                    unit_price_field = get_obj_field("UnitPrice") or get_obj_field("Unit Price")
                    amount_field = get_obj_field("Amount")
                    if isinstance(amount_field, dict) and "amount" in amount_field:
                        amount_val = to_number(amount_field["amount"])
                    else:
                        amount_val = to_number(amount_field)
                    unit_price_val = None
                    if isinstance(unit_price_field, dict) and "amount" in unit_price_field:
                        unit_price_val = to_number(unit_price_field["amount"])
                    else:
                        unit_price_val = to_number(unit_price_field)
                    items.append({
                        "ProductCode": prod_code,
                        "Description": desc,
                        "Quantity": to_number(qty),
                        "UnitPrice": unit_price_val,
                        "Amount": amount_val
                    })

    return {
        "PurchaseOrder": pdf_po,
        "InvoiceId": pdf_inv_id,
        "InvoiceTotal": pdf_invoice_total,
        "Items": items
    }

# ---------- Run ----------
if run_button:
    # server-side credential check
    if not AZURE_ENDPOINT or not AZURE_KEY:
        st.error("Server-side Azure credentials missing. Set AZURE_ENDPOINT and AZURE_KEY as environment variables.")
        st.stop()

    if uploaded_pdf is None or uploaded_excel is None:
        st.error("Please upload both PDF and Excel files.")
        st.stop()

    # READ EXCEL and apply header row
    try:
        temp_df = pd.read_excel(uploaded_excel, header=None, engine="openpyxl")
    except Exception as e:
        st.error(f"Failed to read uploaded Excel: {e}")
        st.stop()

    if temp_df.shape[0] <= header_row_idx:
        st.error(f"Excel has fewer than {header_row_idx+1} rows; cannot use row {header_row_idx} as header.")
        st.stop()

    new_header = temp_df.iloc[header_row_idx].tolist()
    df = temp_df.iloc[header_row_idx+1:].copy().reset_index(drop=True)
    df.columns = [str(h).strip() for h in new_header]
    df = df.dropna(axis=1, how="all")
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)

    # normalize column names
    if "PO Ref No" not in df.columns:
        for c in df.columns:
            if "po" in str(c).lower():
                df.rename(columns={c: "PO Ref No"}, inplace=True)
                break
    if "SKU" not in df.columns:
        for c in df.columns:
            if "sku" in str(c).lower():
                df.rename(columns={c: "SKU"}, inplace=True)
                break
    if "Name" not in df.columns:
        for c in df.columns:
            if "name" in str(c).lower():
                df.rename(columns={c: "Name"}, inplace=True)
                break

    # numeric columns
    def create_numeric_if_missing(orig_col_name, target_col_name):
        if target_col_name in df.columns:
            return
        norm_map = {col.lower().replace(" ", ""): col for col in df.columns}
        key = orig_col_name.lower().replace(" ", "")
        if key in norm_map:
            df[target_col_name] = df[norm_map[key]].apply(to_number)
        else:
            if orig_col_name in df.columns:
                df[target_col_name] = df[orig_col_name].apply(to_number)

    create_numeric_if_missing("Total Value", "Total Value_num")
    create_numeric_if_missing("Included Tax", "Included Tax_num")
    create_numeric_if_missing("Base Price", "Base Price_num")
    create_numeric_if_missing("Item Price", "Item Price_num")
    if "Ordered Quantity" in df.columns:
        df["Ordered Quantity"] = df["Ordered Quantity"].apply(lambda x: int(to_number(x)) if to_number(x) is not None else None)

    if "SKU" not in df.columns:
        st.error("Could not detect SKU column in the Excel. Please ensure a SKU column exists.")
        st.stop()

    df["SKU_norm"] = df["SKU"].fillna("").apply(lambda x: normalize_sku_for_match(x))

    # Azure client
    try:
        client = DocumentIntelligenceClient(endpoint=AZURE_ENDPOINT, credential=AzureKeyCredential(AZURE_KEY))
    except Exception as e:
        st.error(f"Failed to create Azure client: {e}")
        st.stop()

    # parse pdf
    try:
        pdf_bytes = uploaded_pdf.read()
        pdf_extracted = parse_invoice_from_bytes(client, pdf_bytes)
        if pdf_extracted is None:
            st.error("No document extracted from PDF via Azure.")
            st.stop()
    except Exception as e:
        st.error(f"Failed to extract invoice from PDF: {e}")
        st.stop()

    # Show summary cards
    po_pdf = str(pdf_extracted.get("PurchaseOrder")).strip() if pdf_extracted.get("PurchaseOrder") else None
    pdf_inv_id = pdf_extracted.get("InvoiceId")
    pdf_total = pdf_extracted.get("InvoiceTotal")
    po_exists = False
    if po_pdf and "PO Ref No" in df.columns:
        po_exists = df["PO Ref No"].astype(str).str.strip().str.lower().eq(po_pdf.lower()).any()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Invoice PO", value=po_pdf if po_pdf else "—")
    c2.metric("Invoice ID", value=pdf_inv_id if pdf_inv_id else "—")
    c3.metric("Invoice Total (PDF)", value=f"{pdf_total:.2f}" if pdf_total is not None else "—")
    c4.metric("PO in Excel", value="Yes" if po_exists else "No")

    st.markdown("---")
    st.subheader("Invoice Items (extracted)")
    st.write(pd.DataFrame(pdf_extracted.get("Items", [])))

    if not po_exists:
        st.warning("PurchaseOrder not found in Excel. Matching will be attempted but results may be incomplete.")

    # filter excel rows for PO if exists
    df_po = df
    if po_exists:
        df_po = df[df["PO Ref No"].astype(str).str.strip().str.lower() == po_pdf.lower()].copy()

    # prepare aggregated excel map
    df_po["Total Value_num_f"] = df_po.get("Total Value_num", pd.Series([None]*len(df_po))).apply(lambda x: to_number(x) if x is not None else None)
    df_po["Included Tax_num_f"] = df_po.get("Included Tax_num", pd.Series([0.0]*len(df_po))).apply(lambda x: to_number(x) if x is not None else 0.0)
    df_po["Ordered Quantity_f"] = df_po.get("Ordered Quantity", pd.Series([None]*len(df_po))).apply(lambda x: int(to_number(x)) if to_number(x) is not None else None)
    df_po["SKU_norm"] = df_po["SKU"].fillna("").apply(lambda x: normalize_sku_for_match(x))

    agg = df_po.groupby("SKU_norm").agg({
        "SKU": lambda s: s.iloc[0],
        "Name": (lambda s: s.iloc[0]) if "Name" in df_po.columns else (lambda s: None),
        "Ordered Quantity_f": "sum",
        "Total Value_num_f": "sum",
        "Included Tax_num_f": "sum"
    }).reset_index().rename(columns={
        "Ordered Quantity_f": "Excel_Qty_sum",
        "Total Value_num_f": "Excel_TotalValue_sum",
        "Included Tax_num_f": "Excel_IncTax_sum"
    })

    excel_map = {}
    for _, r in agg.iterrows():
        sku_norm = r["SKU_norm"]
        excel_map[sku_norm] = {
            "SKU": r["SKU"],
            "Name": r.get("Name"),
            "Excel_Qty_sum": r["Excel_Qty_sum"] if not pd.isna(r["Excel_Qty_sum"]) else None,
            "Excel_TotalValue_sum": r["Excel_TotalValue_sum"] if not pd.isna(r["Excel_TotalValue_sum"]) else None,
            "Excel_IncTax_sum": r["Excel_IncTax_sum"] if not pd.isna(r["Excel_IncTax_sum"]) else 0.0
        }

    # compare items
    recon_rows = []
    qty_matches = 0
    amt_matches = 0
    pdf_items_amount_sum = 0.0
    items = pdf_extracted.get("Items", [])

    for item in items:
        pdf_code = str(item.get("ProductCode")) if item.get("ProductCode") else ""
        pdf_qty = int(item.get("Quantity")) if item.get("Quantity") is not None else None
        pdf_line_amount = to_number(item.get("Amount"))
        pdf_items_amount_sum += (pdf_line_amount or 0.0)
        pdf_sku_norm = normalize_sku_for_match(pdf_code)

        excel_entry = excel_map.get(pdf_sku_norm)
        if excel_entry is None:
            # fallback contains
            found = None
            for k in excel_map.keys():
                if pdf_sku_norm in k or k in pdf_sku_norm:
                    found = k; break
            if found:
                excel_entry = excel_map[found]

        if excel_entry is None:
            recon_rows.append({
                "PDF SKU": pdf_code,
                "PDF Qty": pdf_qty,
                "Excel SKU": None,
                "Excel Qty (sum)": None,
                "Qty Match": False,
                "PDF Line Amount": pdf_line_amount,
                "Excel Product Amount (Total - IncTax)": None,
                "Excel Included Tax (sum)": None,
                "Amount Match": False,
                "Notes": "SKU NOT FOUND"
            })
            continue

        excel_qty_sum = excel_entry.get("Excel_Qty_sum")
        excel_total_sum = excel_entry.get("Excel_TotalValue_sum")
        excel_inc_tax_sum = excel_entry.get("Excel_IncTax_sum") or 0.0
        excel_product_amount = None
        if excel_total_sum is not None:
            excel_product_amount = (excel_total_sum - excel_inc_tax_sum) if excel_inc_tax_sum is not None else excel_total_sum

        qty_match = (excel_qty_sum is not None and pdf_qty is not None and int(excel_qty_sum) == int(pdf_qty))
        amount_match = False
        if pdf_line_amount is not None and excel_product_amount is not None:
            amount_match = amounts_close(pdf_line_amount, excel_product_amount, abs_tol=abs_tol, rel_tol=rel_tol)
        elif pdf_line_amount is not None and excel_total_sum is not None:
            amount_match = amounts_close(pdf_line_amount, excel_total_sum, abs_tol=abs_tol, rel_tol=rel_tol)

        if qty_match: qty_matches += 1
        if amount_match: amt_matches += 1

        recon_rows.append({
            "PDF SKU": pdf_code,
            "Excel SKU": excel_entry.get("SKU"),
            "PDF Qty": pdf_qty,
            "Excel Qty (sum)": excel_qty_sum,
            "Qty Match": qty_match,
            "PDF Line Amount": pdf_line_amount,
            "Excel Product Amount (Total - IncTax)": excel_product_amount,
            "Excel Included Tax (sum)": excel_inc_tax_sum,
            "Amount Match": amount_match,
            "Notes": None
        })

    recon_df = pd.DataFrame(recon_rows)
    # optionally filter only mismatches
    if show_only_mismatches and not recon_df.empty:
        recon_df = recon_df[~(recon_df["Qty Match"] & recon_df["Amount Match"])]

    # grand totals
    excel_grand_total = df_po["Total Value_num"].apply(lambda x: to_number(x) if x is not None else 0.0).sum()
    pdf_grand_total = pdf_extracted.get("InvoiceTotal") if pdf_extracted.get("InvoiceTotal") is not None else pdf_items_amount_sum
    grand_total_match = amounts_close(pdf_grand_total, excel_grand_total, abs_tol=abs_tol, rel_tol=rel_tol)

    # show tables with nice layout
    st.subheader("Per-line reconciliation")
    if recon_df.empty:
        st.info("No per-line matches to show (maybe no items extracted or PO not found).")
    else:
        st.dataframe(recon_df)

    # top-level metrics
    pdf_item_count = len(items)
    qty_accuracy_pct = (qty_matches / pdf_item_count * 100) if pdf_item_count else 0.0
    amount_accuracy_pct = (amt_matches / pdf_item_count * 100) if pdf_item_count else 0.0
    grand_score = 100.0 if grand_total_match else max(0.0, 100.0 - (abs(pdf_grand_total - excel_grand_total) / max(abs(excel_grand_total), 1e-9) * 100))
    composite_accuracy = (0.40 * grand_score) + (0.30 * qty_accuracy_pct) + (0.30 * amount_accuracy_pct)

    st.subheader("Summary metrics")
    metrics_cols = st.columns(4)
    metrics_cols[0].metric("PDF lines", pdf_item_count)
    metrics_cols[1].metric("Qty match %", f"{qty_accuracy_pct:.2f}%")
    metrics_cols[2].metric("Line-amount match %", f"{amount_accuracy_pct:.2f}%")
    metrics_cols[3].metric("Composite accuracy", f"{composite_accuracy:.2f}%")
    st.write({
        "PDF grand total": pdf_grand_total,
        "Excel grand total": excel_grand_total,
        "Grand total match": grand_total_match
    })

    # ---------- Prepare styled Excel for download ----------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Excel rows used
        df_po.to_excel(writer, sheet_name="Excel_rows_for_PO", index=False)
        # Per line
        recon_df.to_excel(writer, sheet_name="Per_line_recon", index=False)
        # Summary metrics (single row)
        metrics_df = pd.DataFrame([{
            "PDF filename": getattr(uploaded_pdf, "name", "uploaded.pdf"),
            "PO (PDF)": po_pdf,
            "PO found in Excel": po_exists,
            "Invoice ID (PDF)": pdf_inv_id,
            "PDF grand total": pdf_grand_total,
            "Excel grand total": excel_grand_total,
            "Grand total match": grand_total_match,
            "Qty matches": qty_matches,
            "Line amount matches": amt_matches,
            "Composite accuracy (%)": composite_accuracy
        }])
        metrics_df.to_excel(writer, sheet_name="Summary_metrics", index=False)

    # Style workbook via openpyxl
    output.seek(0)
    wb = load_workbook(output)
    # style headers for each sheet
    header_fill = PatternFill(start_color="FFD9E1", end_color="FFD9E1", fill_type="solid")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # style header row
        for col in range(1, ws.max_column+1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center
            # adjust col width
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = max(12, min(40, len(str(cell.value)) + 5))

    # color Per_line_recon rows based on matches
    if "Per_line_recon" in wb.sheetnames:
        ws = wb["Per_line_recon"]
        # find column indices for helpful columns
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = {name: idx+1 for idx, name in enumerate(header)}
        for r in range(2, ws.max_row+1):
            qty_ok = False
            amt_ok = False
            notes = None
            if "Qty Match" in col_idx:
                v = ws.cell(row=r, column=col_idx["Qty Match"]).value
                qty_ok = bool(v) if v is not None else False
            if "Amount Match" in col_idx:
                v = ws.cell(row=r, column=col_idx["Amount Match"]).value
                amt_ok = bool(v) if v is not None else False
            if "Notes" in col_idx:
                notes = ws.cell(row=r, column=col_idx["Notes"]).value

            # Decide fill
            if notes and "NOT FOUND" in str(notes).upper():
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
            elif qty_ok and amt_ok:
                fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")  # green-ish
            else:
                fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")  # red-ish

            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = fill

    # style Summary_metrics sheet better
    if "Summary_metrics" in wb.sheetnames:
        ws = wb["Summary_metrics"]
        # widen columns
        for col in range(1, ws.max_column+1):
            ws.column_dimensions[get_column_letter(col)].width = max(15, ws.column_dimensions[get_column_letter(col)].width)
        # bold header already done

    # Prepare download
    out_io = io.BytesIO()
    wb.save(out_io)
    out_io.seek(0)
    
    st.download_button(
        label="📥 Download Styled Reconciliation Report",
        data=out_io,
        file_name=get_download_filename(f"Hafele_Reconciliation_{pdf_inv_id}"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.success("Reconciliation complete — improved UI + styled Excel ready for download.")

    # Save to MongoDB
    try:
         save_reconciliation_report(
            collection_name="hafele_reconciliation",
            invoice_no=pdf_inv_id,
            summary_data=metrics_df,
            line_items_data=recon_df,
            metadata={
                "composite_accuracy": composite_accuracy,
                "file_name_pdf": getattr(uploaded_pdf, "name", "uploaded.pdf"),
                "timestamp": str(pd.Timestamp.now())
            }
        )
    except Exception as e:
        st.error(f"Failed to auto-save to MongoDB: {e}")

